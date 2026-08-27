#!/usr/bin/env python3

from __future__ import annotations

import importlib.util
from pathlib import Path
import tempfile
import unittest

from openpyxl import Workbook


ROOT = Path(__file__).resolve().parents[1]
SCRIPT = ROOT / 'scripts' / 'sync_cosmo_weekly.py'
SPEC = importlib.util.spec_from_file_location('sync_cosmo_weekly', SCRIPT)
if SPEC is None or SPEC.loader is None:
    raise RuntimeError('COSMO 동기화 모듈을 불러올 수 없습니다')
MODULE = importlib.util.module_from_spec(SPEC)
SPEC.loader.exec_module(MODULE)


ANCHORS = {
    '영업현황': {'B14': '합계', 'C6': '리테일', 'C7': '캐터링', 'C8': '파우치'},
    '판매현황': {'C6': 'RETAIL', 'C7': 'CATERING', 'C8': '파우치', 'C20': '수출', 'C21': '내수'},
    '생산현황': {'B6': 'CBU', 'B8': 'FBU'},
    '원어구매현황': {'C22': '합계'},
    '재고현황': {'C9': '소계', 'C14': '소계', 'B27': '합계'},
    '자금현황': {'B29': '합계'},
}


def workbook(path: Path, *, drift: bool = False, formula_error: bool = False) -> Path:
    book = Workbook()
    book.remove(book.active)
    for sheet_name in MODULE.EXPECTED_SHEETS:
        book.create_sheet(sheet_name)
    for sheet_name, anchors in ANCHORS.items():
        for coord, value in anchors.items():
            book[sheet_name][coord] = value
    if drift:
        book['영업현황']['C6'] = '리테일 이동'
    if formula_error:
        book['Sheet1']['A1'] = '#REF!'
        book['Sheet1']['A1'].data_type = 'e'
    book.save(path)
    return path


class CosmoWeeklySyncTest(unittest.TestCase):
    def test_semantic_anchors_accept_the_current_template(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            path = workbook(Path(temporary) / 'COSMO 주간보고 (34주차)-첨부파일.xlsx')
            source = MODULE.CosmoWorkbook(path)
            MODULE.validate_semantic_anchors(source)

    def test_semantic_anchor_drift_is_rejected(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            path = workbook(
                Path(temporary) / 'COSMO 주간보고 (34주차)-첨부파일.xlsx',
                drift=True,
            )
            source = MODULE.CosmoWorkbook(path)
            with self.assertRaisesRegex(ValueError, '영업현황!C6'):
                MODULE.validate_semantic_anchors(source)

    def test_formula_errors_are_rejected_before_extraction(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            path = workbook(
                Path(temporary) / 'COSMO 주간보고 (34주차)-첨부파일.xlsx',
                formula_error=True,
            )
            with self.assertRaisesRegex(ValueError, 'Sheet1!A1'):
                MODULE.CosmoWorkbook(path)


if __name__ == '__main__':
    unittest.main()
