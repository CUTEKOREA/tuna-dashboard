#!/usr/bin/env python3
"""파노피 2025 확정 재무제표 추출 (신라교역 회계팀, 2025-12-31).

원본: Financial Statements(2025.12.31)_260119 - 신라교역 회계팀.xlsx
  IS 시트  — 손익계산서 (GHC·USD 병기, 2025/2024)
  FP 시트  — 재무상태표 (동일)
주의: Google Drive 스트리밍 파일 직접 읽기 금지 (코스모 회귀 사례) — 로컬 사본 경로를 인자로.

  python3 scripts/extract_panofi_fs.py <로컬사본.xlsx>
"""
from __future__ import annotations

import json
import sys

import openpyxl

OUT = 'public/data/panofi/panofi_fs_2025.json'

IS_ROWS = {
    6: 'revenue', 9: 'cos', 18: 'gp', 19: 'sga', 20: 'op',
    22: 'interestIncome', 25: 'assetDisposalGain', 26: 'fxTranslationGain', 27: 'fxTradeGain',
    29: 'otherIncome', 31: 'badDebt', 32: 'interestExpense',
    35: 'fxTranslationLoss', 36: 'fxTradeLoss', 37: 'otherExpense',
    38: 'pretax', 39: 'tax', 40: 'net',
    42: 'fxTotal', 44: 'netExFx', 45: 'netExFxExDisposal', 48: 'perVessel',
}
FP_ROWS = {
    6: 'currentAssets', 7: 'cash', 8: 'receivables', 13: 'fishInventory',
    15: 'nonCurrentAssets', 17: 'vessels', 18: 'vesselDep',
    28: 'totalAssets', 29: 'currentLiabilities', 30: 'payables', 31: 'accruedPayables',
    35: 'shortTermDebt', 36: 'nonCurrentLiabilities', 37: 'longTermFxPayable', 38: 'longTermDebt',
    39: 'totalLiabilities', 42: 'totalEquity',
}


def main() -> int:
    src = sys.argv[1]
    wb = openpyxl.load_workbook(src, data_only=True)

    def num(v):
        return round(float(v), 2) if isinstance(v, (int, float)) else None

    is_ws = wb['IS']
    fp_ws = wb['FP']
    payload = {
        '_meta': {
            'source': 'Financial Statements(2025.12.31)_260119 - 신라교역 회계팀.xlsx',
            'asOf': '2025-12-31',
            'basis': '회계 확정 결산 (GHC 장부 → USD 환산) — 판매원장·전략보고와 축이 다르다',
            'generator': 'scripts/extract_panofi_fs.py',
        },
        'fx': {
            # IS 헤더의 평균환율, FP 헤더의 기말환율 (GHC/USD)
            'avg2025': num(is_ws['G3'].value),
            'avg2024': num(is_ws['K3'].value),
            'close2025': num(fp_ws['D3'].value),
            'close2024': num(fp_ws['F3'].value),
        },
        'income': {}, 'position': {},
    }
    for r, key in IS_ROWS.items():
        payload['income'][key] = {'y2025': num(is_ws.cell(r, 7).value), 'y2024': num(is_ws.cell(r, 11).value)}
    for r, key in FP_ROWS.items():
        payload['position'][key] = {'y2025': num(fp_ws.cell(r, 4).value), 'y2024': num(fp_ws.cell(r, 6).value)}

    # 원자료 결함 교정 (검증됨): FP D15(2025 비유동자산 USD)가 $50.3M로 적혀 있으나
    # GHC 240.86M ÷ 기말환율 10.45 = $23.05M = 구성 항목 합과 일치. USD 환산 열의 오기라
    # 파생값으로 교정하고 보고값은 asReported로 보존한다 (유동+비유동=총자산 항등식 회복).
    nca = payload['position']['nonCurrentAssets']
    ghc_nca = num(fp_ws.cell(15, 3).value)
    derived = round(ghc_nca / payload['fx']['close2025'], 2)
    if abs(nca['y2025'] - derived) > 1000:
        payload['_meta']['sourceDefects'] = [
            f"FP D15 비유동자산 USD 보고값 {nca['y2025']:,} ≠ GHC 환산 {derived:,} (구성 합 일치) — 환산 열 오기로 판단, 파생값 채택",
        ]
        nca['y2025AsReported'] = nca['y2025']
        nca['y2025'] = derived

    # 항등식 검산: 유동+비유동 = 총자산
    for yr in ('y2025', 'y2024'):
        pp = payload['position']
        assert abs((pp['currentAssets'][yr] + pp['nonCurrentAssets'][yr]) - pp['totalAssets'][yr]) < 1500, ('CA+NCA', yr)

    # 무결성 검산 — 자산 = 부채 + 자본, 외환합계·제외손익이 구성과 일치
    p, i = payload['position'], payload['income']
    for yr in ('y2025', 'y2024'):
        assert abs(p['totalAssets'][yr] - (p['totalLiabilities'][yr] + p['totalEquity'][yr])) < 2, yr
        fx = (i['fxTranslationGain'][yr] + i['fxTradeGain'][yr]
              + i['fxTranslationLoss'][yr] + i['fxTradeLoss'][yr])
        assert abs(fx - i['fxTotal'][yr]) < 2, ('fx', yr)
        assert abs((i['net'][yr] - i['fxTotal'][yr]) - i['netExFx'][yr]) < 2, ('netExFx', yr)

    with open(OUT, 'w', encoding='utf-8') as fh:
        json.dump(payload, fh, ensure_ascii=False, indent=1)
    print(f"→ {OUT} · net25 {i['net']['y2025']:,} · netExFx25 {i['netExFx']['y2025']:,} · equity25 {p['totalEquity']['y2025']:,}")
    return 0


if __name__ == '__main__':
    raise SystemExit(main())
