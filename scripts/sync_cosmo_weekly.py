#!/usr/bin/env python3
"""Append or replace one COSMO weekly workbook in the dashboard snapshot."""

from __future__ import annotations

import argparse
from datetime import datetime
import hashlib
import json
import math
from pathlib import Path
import re
from typing import Any

from openpyxl import load_workbook


EXPECTED_SHEETS = [
    '영업현황',
    '판매현황',
    '생산현황',
    '원어구매현황',
    '재고현황',
    '자금현황',
    'Sheet1',
]
SEMANTIC_ANCHORS = {
    '영업현황': {'B14': '합계', 'C6': '리테일', 'C7': '캐터링', 'C8': '파우치'},
    '판매현황': {'C6': 'RETAIL', 'C7': 'CATERING', 'C8': '파우치', 'C20': '수출', 'C21': '내수'},
    '생산현황': {'B6': 'CBU', 'B8': 'FBU'},
    '원어구매현황': {'C22': '합계'},
    '재고현황': {'C9': '소계', 'C14': '소계', 'B27': '합계'},
    '자금현황': {'B29': '합계'},
}
WEEK_RE = re.compile(r'\((\d+)주차\)')
PERIOD_RE = re.compile(r'(\d{1,2}/\d{1,2})')
ERROR_VALUES = {'#REF!', '#DIV/0!', '#VALUE!', '#N/A', '#NAME?', '#NUM!', '#NULL!'}


def number(value: Any) -> float | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        result = float(value)
        return result if math.isfinite(result) else None
    if isinstance(value, str):
        normalized = value.strip().replace(',', '')
        if not normalized or normalized in {'-', 'N/A'}:
            return None
        try:
            result = float(normalized)
        except ValueError:
            return None
        return result if math.isfinite(result) else None
    return None


def rounded(value: float) -> float:
    result = round(value, 2)
    return 0.0 if result == 0 else result


class CosmoWorkbook:
    def __init__(self, path: Path) -> None:
        self.path = path
        self.formulas = load_workbook(path, data_only=False, read_only=False)
        self.values = load_workbook(path, data_only=True, read_only=False)
        if self.formulas.sheetnames != EXPECTED_SHEETS:
            raise ValueError(
                f'예상 시트 {EXPECTED_SHEETS!r}와 다릅니다: {self.formulas.sheetnames!r}',
            )
        self._assert_no_formula_errors()

    def _assert_no_formula_errors(self) -> None:
        errors: list[str] = []
        for sheet in self.values.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    value = cell.value
                    if cell.data_type == 'e' or (isinstance(value, str) and value in ERROR_VALUES):
                        errors.append(f'{sheet.title}!{cell.coordinate}={value}')
        if errors:
            raise ValueError('엑셀 수식 오류: ' + ', '.join(errors))

    def value(self, sheet: str, coord: str, *, required: bool = False) -> Any:
        formula_cell = self.formulas[sheet][coord]
        value = self.values[sheet][coord].value
        is_formula = formula_cell.data_type == 'f' or (
            isinstance(formula_cell.value, str) and formula_cell.value.startswith('=')
        )
        if required and value is None:
            reason = '수식 캐시 누락' if is_formula else '필수 값 누락'
            raise ValueError(f'{reason}: {sheet}!{coord}')
        return value

    def num(self, sheet: str, coord: str, *, required: bool = False) -> float | None:
        value = self.value(sheet, coord, required=required)
        result = number(value)
        if required and result is None:
            raise ValueError(f'필수 숫자 아님: {sheet}!{coord}={value!r}')
        return result


def validate_semantic_anchors(book: CosmoWorkbook) -> None:
    for sheet, anchors in SEMANTIC_ANCHORS.items():
        for coord, expected in anchors.items():
            actual = book.value(sheet, coord, required=True)
            if actual != expected:
                raise ValueError(
                    f'템플릿 의미 라벨 불일치: {sheet}!{coord}={actual!r}, 예상 {expected!r}',
                )


def period(value: Any, coord: str) -> str:
    match = PERIOD_RE.search(str(value))
    if not match:
        raise ValueError(f'기간을 읽을 수 없습니다: {coord}={value!r}')
    return match.group(1)


def parse_week(path: Path) -> tuple[dict[str, Any], list[dict[str, Any]]]:
    match = WEEK_RE.search(path.name)
    if not match:
        raise ValueError(f'파일명에서 주차를 찾을 수 없습니다: {path.name}')
    week = int(match.group(1))
    book = CosmoWorkbook(path)
    title = str(book.value('영업현황', 'B2', required=True))
    if f'{week}주차' not in title:
        raise ValueError(f'파일명 주차와 영업현황 제목이 다릅니다: {path.name} / {title}')
    validate_semantic_anchors(book)

    backlog_rows = [
        (6, 'RETAIL', '리테일'),
        (7, 'CATERING', '캐터링'),
        (8, 'POUCH', '파우치'),
    ]
    backlog_items = [
        {
            'item': item,
            'label': label,
            'qty': book.num('영업현황', f'H{row}'),
            'unitPrice': book.num('영업현황', f'I{row}'),
            'usd': book.num('영업현황', f'J{row}'),
        }
        for row, item, label in backlog_rows
    ]
    new_orders = [
        {
            'item': item,
            'label': label,
            'qty': book.num('영업현황', f'E{row}'),
            'usd': book.num('영업현황', f'G{row}'),
        }
        for row, item, label in backlog_rows
        if book.num('영업현황', f'E{row}') not in (None, 0)
    ]

    sales_rows = [
        (6, 'RETAIL', '리테일'),
        (7, 'CATERING', '캐터링'),
        (8, 'POUCH', '파우치'),
        (9, 'COOKLOIN', '쿡로인'),
        (10, 'DOMESTIC_CAN', '참치캔(내수)'),
        (11, 'FISHMEAL_EXPORT', '피쉬밀 수출'),
        (12, 'FISHMEAL_LOCAL', '피쉬밀 로컬'),
        (13, 'FISHHEAD', '피쉬헤드'),
        (14, 'FISHWASTE', '피쉬 WASTE'),
        (17, 'FROZEN_LOIN', '냉동로인'),
        (18, 'RAW_FISH', '원어(FBU)'),
    ]
    sales = [
        {
            'item': item,
            'label': label,
            'unit': book.value('판매현황', f'D{row}', required=True),
            'weekQty': book.num('판매현황', f'E{row}'),
            'weekPrice': book.num('판매현황', f'F{row}'),
            'weekUsd': book.num('판매현황', f'G{row}'),
            'cumQty': book.num('판매현황', f'H{row}'),
            'cumPrice': book.num('판매현황', f'I{row}'),
            'cumUsd': book.num('판매현황', f'J{row}'),
        }
        for row, item, label in sales_rows
    ]

    def production(row: int, *, planned: bool) -> dict[str, float | None]:
        result = {
            'weekDays': book.num('생산현황', f'C{row}'),
            'weekRawMt': book.num('생산현황', f'D{row}'),
            'weekYield': book.num('생산현황', f'E{row}'),
            'weekDaily': book.num('생산현황', f'F{row}'),
            'cumDays': book.num('생산현황', f'G{row}'),
            'cumRawMt': book.num('생산현황', f'H{row}'),
            'cumYield': book.num('생산현황', f'I{row}'),
            'cumDaily': book.num('생산현황', f'J{row}'),
            'planDays': None,
            'planRawMt': None,
            'planYield': None,
            'planDaily': None,
            'gapDays': None,
            'gapRawMt': None,
            'gapYield': None,
            'gapDaily': None,
        }
        if planned:
            result.update({
                'planDays': book.num('생산현황', f'P{row}'),
                'planRawMt': book.num('생산현황', f'Q{row}'),
                'planYield': book.num('생산현황', f'R{row}'),
                'planDaily': book.num('생산현황', f'S{row}'),
                'gapDays': book.num('생산현황', f'K{row}'),
                'gapRawMt': book.num('생산현황', f'L{row}'),
                'gapYield': book.num('생산현황', f'M{row}'),
                'gapDaily': book.num('생산현황', f'N{row}'),
            })
        return result

    purchase_rows = [
        (6, 'PS', '파노피', 'SJ'),
        (7, 'PS', '파노피', 'YF/BE'),
        (9, 'PS', '타선사', 'SJ'),
        (10, 'PS', '타선사', 'YF/BE'),
        (13, 'FBU', '파노피', 'YF'),
        (14, 'FBU', '파노피', 'BE'),
        (16, 'FBU', '기타선사', 'YF'),
        (17, 'FBU', '기타선사', 'BE'),
    ]
    purchase_lines = [
        {
            'unit': unit,
            'supplier': supplier,
            'species': species,
            'weekMt': book.num('원어구매현황', f'E{row}'),
            'weekPrice': book.num('원어구매현황', f'F{row}'),
            'cumMt': book.num('원어구매현황', f'H{row}'),
            'cumPrice': book.num('원어구매현황', f'I{row}'),
        }
        for row, unit, supplier, species in purchase_rows
    ]

    inventory_rows = [
        (6, '원어'), (7, '원어'), (8, '원어'),
        (10, '제품(CBU)'), (11, '제품(CBU)'), (12, '제품(CBU)'),
        (13, '제품(FBU)'),
        (16, '공관'), (17, '공관'), (18, '공관'),
        (20, 'ENDS'), (21, 'ENDS'),
        (23, '주입액'), (24, '주입액'), (25, '주입액'),
    ]
    inventory_lines = [
        {
            'group': group,
            'item': book.value('재고현황', f'C{row}', required=True),
            'unit': book.value('재고현황', f'D{row}', required=True),
            'beginQty': book.num('재고현황', f'E{row}'),
            'beginUsd': book.num('재고현황', f'G{row}'),
            'inQty': book.num('재고현황', f'H{row}'),
            'outQty': book.num('재고현황', f'K{row}'),
            'endQty': book.num('재고현황', f'N{row}'),
            'endUsd': book.num('재고현황', f'P{row}'),
        }
        for row, group in inventory_rows
    ]

    currency_rows = [(24, 'USD'), (25, 'GHC'), (26, 'EUR'), (27, 'GBP'), (28, 'JPY')]
    cash_by_currency = [
        {
            'ccy': currency,
            'beginUsd': book.num('자금현황', f'G{row}'),
            'endUsd': book.num('자금현황', f'P{row}'),
        }
        for row, currency in currency_rows
    ]

    week_data: dict[str, Any] = {
        'source': path.name,
        'sha256': hashlib.sha256(path.read_bytes()).hexdigest()[:16],
        'year': 2026,
        'week': week,
        'backlog_total_fcl': book.num('영업현황', 'H14', required=True),
        'backlog_total_usd': book.num('영업현황', 'J14', required=True),
        'new_orders_fcl': book.num('영업현황', 'E14', required=True),
        'new_orders_usd': book.num('영업현황', 'G14', required=True),
        'backlogItems': backlog_items,
        'newOrders': new_orders,
        'exportWeekUsd': book.num('판매현황', 'G20', required=True),
        'exportCumUsd': book.num('판매현황', 'J20', required=True),
        'domesticWeekUsd': book.num('판매현황', 'G21', required=True),
        'domesticCumUsd': book.num('판매현황', 'J21', required=True),
        'sales': sales,
        'salesWeekUsd': book.num('판매현황', 'G22', required=True),
        'salesCumUsd': book.num('판매현황', 'J22', required=True),
        'production': {
            'CBU': production(6, planned=True),
            'FBU': production(8, planned=False),
        },
        'purchase': {
            'lines': purchase_lines,
            'panofiCumMt': book.num('원어구매현황', 'H20', required=True),
            'weekMt': book.num('원어구매현황', 'E22', required=True),
            'weekUsd': book.num('원어구매현황', 'G22', required=True),
            'cumMt': book.num('원어구매현황', 'H22', required=True),
            'cumUnit': book.num('원어구매현황', 'I22', required=True),
            'cumUsd': book.num('원어구매현황', 'J22', required=True),
        },
        'periodStart': period(book.value('재고현황', 'E4', required=True), '재고현황!E4'),
        'periodEnd': period(book.value('재고현황', 'N4', required=True), '재고현황!N4'),
        'inventory': {
            'lines': inventory_lines,
            'totalBeginUsd': book.num('재고현황', 'G27', required=True),
            'totalInUsd': book.num('재고현황', 'J27', required=True),
            'totalOutUsd': book.num('재고현황', 'M27', required=True),
            'totalEndUsd': book.num('재고현황', 'P27', required=True),
        },
        'ghcRate': book.num('자금현황', 'O8', required=True),
        'cash': {
            'byCurrency': cash_by_currency,
            'beginUsd': book.num('자금현황', 'G29', required=True),
            'inUsd': book.num('자금현황', 'J29', required=True),
            'outUsd': book.num('자금현황', 'M29', required=True),
            'endUsd': book.num('자금현황', 'P29', required=True),
            'transferUsd': book.num('자금현황', 'J35', required=True),
            'externalInUsd': book.num('자금현황', 'J37', required=True),
            'externalOutUsd': book.num('자금현황', 'M37', required=True),
        },
    }

    quotes: list[dict[str, Any]] = []
    quote_sheet = book.values['영업현황']
    for row in range(20, quote_sheet.max_row + 1):
        kind = book.value('영업현황', f'B{row}')
        customer = book.value('영업현황', f'C{row}')
        if kind not in {'Can', 'Pouch'} or not customer:
            continue
        quotes.append({
            'kind': kind,
            'customer': customer,
            'qty': str(book.value('영업현황', f'D{row}', required=True)),
            'expected': str(book.value('영업현황', f'E{row}', required=True)),
            'fish': str(book.value('영업현황', f'F{row}', required=True)),
            'spec': str(book.value('영업현황', f'G{row}', required=True)),
            'style': str(book.value('영업현황', f'H{row}', required=True)),
            'media': str(book.value('영업현황', f'I{row}', required=True)),
            'mfgCost': book.num('영업현황', f'L{row}'),
            'otherCost': book.num('영업현황', f'M{row}'),
            'totalCost': book.num('영업현황', f'N{row}'),
            'sellPrice': book.num('영업현황', f'O{row}'),
            'margin': book.num('영업현황', f'Q{row}'),
            'week': week,
        })
    return week_data, quotes


def make_checks(current: dict[str, Any], previous: dict[str, Any]) -> list[dict[str, Any]]:
    week = current['week']
    current_inventory = current['inventory']
    previous_inventory = previous['inventory']
    current_cbu = current['production']['CBU']
    previous_cbu = previous['production']['CBU']
    current_fbu = current['production']['FBU']
    previous_fbu = previous['production']['FBU']
    current_cash = current['cash']

    checks = [
        ('재고 항등식', rounded(
            current_inventory['totalBeginUsd'] + current_inventory['totalInUsd']
            - current_inventory['totalOutUsd'] - current_inventory['totalEndUsd']
        ), '기초+입고−출고−잔액'),
        ('재고 이월', rounded(
            current_inventory['totalBeginUsd'] - previous_inventory['totalEndUsd']
        ), '금주기초−전주잔액'),
        ('판매 누적 브릿지', rounded(
            current['salesCumUsd'] - previous['salesCumUsd'] - current['salesWeekUsd']
        ), '전주누적+금주−금주누적'),
        ('CBU 생산 누적 브릿지', rounded(
            current_cbu['cumRawMt'] - previous_cbu['cumRawMt'] - current_cbu['weekRawMt']
        ), ''),
        ('CBU 생산일수 누적 브릿지', rounded(
            current_cbu['cumDays'] - previous_cbu['cumDays'] - current_cbu['weekDays']
        ), '전주누계+금주−금주누계'),
        ('FBU 생산 누적 브릿지', rounded(
            current_fbu['cumRawMt'] - previous_fbu['cumRawMt'] - current_fbu['weekRawMt']
        ), ''),
        ('FBU 생산일수 누적 브릿지', rounded(
            current_fbu['cumDays'] - previous_fbu['cumDays'] - current_fbu['weekDays']
        ), '전주누계+금주−금주누계'),
        ('자금 항등식', rounded(
            current_cash['beginUsd'] + current_cash['inUsd']
            - current_cash['outUsd'] - current_cash['endUsd']
        ), ''),
    ]
    return [
        {
            'week': week,
            'name': name,
            'residual': residual,
            'ok': abs(residual) <= 0.01,
            'note': note,
        }
        for name, residual, note in checks
    ]


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument('workbook', type=Path)
    parser.add_argument(
        '--data',
        type=Path,
        default=Path('public/data/cosmo/cosmo_2026.json'),
    )
    parser.add_argument('--generated-at')
    parser.add_argument('--source-label', default='Google Drive/COSMO 주간보고')
    parser.add_argument('--dry-run', action='store_true')
    args = parser.parse_args()

    workbook = args.workbook.expanduser().resolve()
    data_path = args.data.expanduser().resolve()
    if not workbook.is_file():
        raise FileNotFoundError(workbook)
    if not data_path.is_file():
        raise FileNotFoundError(data_path)

    payload = json.loads(data_path.read_text(encoding='utf-8'))
    week_data, week_quotes = parse_week(workbook)
    week = week_data['week']
    weeks = [item for item in payload['weeks'] if item['week'] != week]
    previous = next((item for item in weeks if item['week'] == week - 1), None)
    if previous is None:
        raise ValueError(f'{week - 1}주차 데이터가 없어 연속 검산을 수행할 수 없습니다')
    weeks.append(week_data)
    weeks.sort(key=lambda item: item['week'])
    payload['weeks'] = weeks

    quotes = [item for item in payload['quotes'] if item['week'] != week]
    quotes.extend(week_quotes)
    quotes.sort(key=lambda item: item['week'])
    payload['quotes'] = quotes

    checks = [item for item in payload['checks'] if item['week'] != week]
    checks.extend(make_checks(week_data, previous))
    checks.sort(key=lambda item: item['week'])
    payload['checks'] = checks

    week_numbers = [item['week'] for item in weeks]
    week_min, week_max = min(week_numbers), max(week_numbers)
    payload['meta'] = {
        **payload['meta'],
        'generated': args.generated_at or datetime.now().astimezone().isoformat(timespec='seconds'),
        'sourceDir': args.source_label,
        'weekCount': len(weeks),
        'weekRange': [week_min, week_max],
        'missingWeeks': sorted(set(range(week_min, week_max + 1)) - set(week_numbers)),
        'monthCount': len(payload['monthly']),
        'quoteCount': len(quotes),
        'checkCount': len(checks),
        'checkFailCount': sum(1 for item in checks if not item['ok']),
    }

    summary = {
        'week': week,
        'source': week_data['source'],
        'sha256': week_data['sha256'],
        'quotes': len(week_quotes),
        'checks': checks[-8:],
        'backlogFcl': week_data['backlog_total_fcl'],
        'salesWeekUsd': week_data['salesWeekUsd'],
    }
    print(json.dumps(summary, ensure_ascii=False, indent=2))
    if not args.dry_run:
        data_path.write_text(
            json.dumps(payload, ensure_ascii=False, separators=(',', ':')),
            encoding='utf-8',
        )
    return 0


if __name__ == '__main__':
    raise SystemExit(main())
