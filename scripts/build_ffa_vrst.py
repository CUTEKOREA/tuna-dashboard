#!/usr/bin/env python3
"""FFA 조업허가 선단(Good Standing) + VMS 주간 보고현황 집계.

원본: FFA VRST 주간 보고서 xlsx (2026-08-01~14).
  시트 3개 — FFAVMS_VRST_Report(선박별 일별 보고건수) / Countbyflag(국기×선종 집계)
            / Good Standing Vessels(제원·선주)

⚠ **원본 집계 시트를 쓰지 않는다.** Countbyflag 의 선종 열 합이 819 인데 실제는 820 이다
  (중국 행: TOTAL 303, 선종합 302). 원표에서 직접 세고, 이 불일치를 _meta 에 남긴다.

⚠ **어창 용량은 단위가 섞여 있다** (Cubic Metres / Metric Tonnes). 더하거나 한 축에
  얹으면 안 된다 — 단위를 함께 내보내고 집계는 하지 않는다.

⚠ **선주명 표기가 흔들린다** ("Silla Co., Ltd" / "Silla Co. Ltd",
  "Dongwon Industries Co., Ltd" / "Co., Ltd." / "CO.,LTD"). 정규화해 묶되
  원문 표기를 함께 남긴다.

실행: python3 scripts/build_ffa_vrst.py <xlsx> [--out public/data]
"""
from __future__ import annotations

import argparse
import json
import re
import statistics
import sys
from collections import Counter, defaultdict
from pathlib import Path

import openpyxl

FLAG_KO = {
    'China': '중국', 'Taiwan': '대만', 'Japan': '일본', 'Panama': '파나마',
    'Korea, Republic of': '한국', 'Philippines': '필리핀', 'Fiji': '피지',
    'Micronesia, Federated States of': '미크로네시아', 'Vanuatu': '바누아투',
    'Nauru': '나우루', 'Kiribati': '키리바시', 'United States': '미국',
    'Marshall Islands': '마셜제도', 'Papua New Guinea': '파푸아뉴기니',
    'Solomon Islands': '솔로몬제도', 'Cook Islands': '쿡제도', 'Ecuador': '에콰도르',
    'Spain': '스페인', 'Tuvalu': '투발루', 'El Salvador': '엘살바도르',
}

# 원표의 선종 표기가 여러 갈래다. 화면에 쓸 한글 한 갈래로 묶는다.
TYPE_KO = {
    'LONGLINE': '연승', 'LONGLINER': '연승', 'TUNA LONGLINER': '연승',
    'SINGLE PURSE SEINER': '선망', 'TUNA PURSE SEINER': '선망', 'PURSE SEINER': '선망',
    'FISH CARRIER/REEFER': '운반선', 'FISH CARRIER': '운반선',
    'BUNKER': '급유선', 'POLE AND LINE': '채낚기',
    'SCOUT VESSEL': '탐색선', 'SEARCH/ANCHOR/LIGHT': '집어·양묘',
}


def norm_company(name: str) -> str:
    """법인 표기 흔들림을 지운다. 대소문자·구두점·법인격 접미만 건드린다."""
    s = re.sub(r'[.,]', ' ', str(name or '')).upper()
    s = re.sub(r'\b(CO|LTD|LIMITED|INC|CORP|CORPORATION|PTE|PTY|LLC)\b', ' ', s)
    return re.sub(r'\s+', ' ', s).strip()


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument('xlsx')
    ap.add_argument('--out', default='public/data')
    args = ap.parse_args()

    wb = openpyxl.load_workbook(args.xlsx, read_only=True, data_only=True)
    vr = list(wb['FFAVMS_VRST_Report'].iter_rows(values_only=True))
    gs = list(wb['Good Standing Vessels'].iter_rows(values_only=True))

    vhdr = vr[0]
    vbody = [r for r in vr[1:] if r[0]]
    # 일별 열은 헤더가 날짜다. 위치를 세지 말고 날짜인 열을 찾는다.
    day_idx = [i for i, h in enumerate(vhdr) if hasattr(h, 'strftime')]
    days = [vhdr[i].strftime('%Y-%m-%d') for i in day_idx]

    ghdr = gs[0]
    G = {h: i for i, h in enumerate(ghdr) if h}
    gbody = [r for r in gs[1:] if r[0]]
    specs = {str(r[0]).strip(): r for r in gbody}

    errors: list[str] = []
    if len(vbody) != len(gbody):
        errors.append(f'시트 행수 불일치: VRST {len(vbody)} vs 제원 {len(gbody)}')

    # ── 원본 집계 시트와 대조 (틀리면 기록하고 원표를 쓴다) ────────────────
    cb = [r for r in wb['Countbyflag'].iter_rows(values_only=True) if r and r[0]]
    chdr = cb[0]
    num = lambda v: v if isinstance(v, (int, float)) else 0
    cb_flag = {r[0]: num(r[1]) for r in cb[1:] if r[0] != 'TOTAL'}
    raw_flag = Counter(r[6] for r in vbody)
    for f, n in raw_flag.items():
        if cb_flag.get(f) != n:
            errors.append(f'국기 집계 불일치 {f}: 원표 {n} vs 집계표 {cb_flag.get(f)}')

    # 집계표 선종 열은 합이 안 맞는다. 어긋난 행을 찾아 기록만 하고 원표를 쓴다.
    type_gap = []
    for r in cb[1:]:
        if r[0] == 'TOTAL':
            continue
        s = sum(num(r[i]) for i in range(2, len(chdr)))
        if s != num(r[1]):
            type_gap.append(f'{FLAG_KO.get(r[0], r[0])} 표기 {num(r[1])}척 · 선종 열 합 {s}척')

    # ── 선박별 VMS 보고 ───────────────────────────────────────────────
    #
    # ⚠ **빈 칸이 두 가지 뜻이다.** 등록일 이후의 빈 칸은 「그날 보고 0건」이고,
    #   등록일 이전의 빈 칸은 「아직 등록 전」이라 셀 대상이 아니다. 둘을 합치면
    #   8월 14일에 등록한 배가 13일치를 무보고한 것처럼 보인다.
    #   그래서 등록일 이전은 분모에서 뺀다.
    vessels, after_period = [], []
    for r in vbody:
        name = str(r[0]).strip()
        sp = specs.get(name)
        reg = sp[G['Registration Start']] if sp and G.get('Registration Start') is not None else None
        reg_day = reg.strftime('%Y-%m-%d') if hasattr(reg, 'strftime') else None

        applicable, nums = [], []
        for d, i in zip(days, day_idx):
            if reg_day and d < reg_day:
                continue  # 등록 전 — 해당 없음
            applicable.append(d)
            v = r[i]
            nums.append(float(v) if isinstance(v, (int, float)) else 0.0)
        if not applicable:
            # 등록일이 보고 기간보다 **뒤**면 실패가 아니다. 명부는 추출 시점 기준이라
            # 월간 보고서에는 월말 직후 등록분이 함께 실린다(2026-08 판에 9/1~9/2 등록 10척).
            # 보고할 날이 없었을 뿐이므로 집계에서 빼고 _meta 에 남긴다.
            # 등록일이 기간보다 **앞**인데 해당일이 없다면 그건 진짜 이상이다.
            if reg_day and reg_day > days[-1]:
                after_period.append({'선박': name, '등록': reg_day})
                continue
            errors.append(f'{name}: 기간 내 등록일이 없다 (등록 {reg_day})')
            continue

        # 정상 주기는 그 배 자신의 중앙값으로 잡는다. 선망 48건/일(30분), 운반선
        # 24건/일(1시간)이 일반이지만 배마다 다를 수 있어 외부 상수를 박지 않는다.
        cadence = statistics.median(nums)
        raw_owner = str(sp[G['Owner Company Name']]).strip() if sp and sp[G['Owner Company Name']] else ''
        # 결손일은 「정상 주기 미달」이되, **무보고일보다 적을 수는 없다.**
        # 절반 넘게 못 받은 배는 중앙값 자체가 0 이라 「미달 0일」로 세어진다 —
        # 그 배야말로 최악인데 무결점으로 뒤집힌다.
        vessels.append({
            'n': name,
            'days적용': len(applicable),
            'f': FLAG_KO.get(r[6], r[6]),
            'g': TYPE_KO.get(r[7], r[7]),
            'mtu': str(r[9]) if r[9] else '',
            'st': '국내' if 'Domestic' in str(r[1]) else '외국',
            'days': nums,
            'cadence': cadence,
            'zero': sum(1 for v in nums if v == 0),
            'short': max(sum(1 for v in nums if v < cadence),
                         sum(1 for v in nums if v == 0)),
            'owner': raw_owner,
            'ownerKey': norm_company(raw_owner),
            'grt': sp[G['Vessel Tonnage (GRT)']] if sp else None,
            'built': sp[G['Built in Year']] if sp else None,
            'imo': str(r[5]) if r[5] else None,
        })

    if len(vessels) + len(after_period) != len(vbody):
        errors.append(f'선박 집계 누락: {len(vessels)}/{len(vbody)} (기간 후 등록 {len(after_period)}척 제외)')

    # ── 선종 표준 주기 ────────────────────────────────────────────────
    #
    # 배 자신의 중앙값만 기준으로 삼으면, **늘 절반만 보내는 배가 만점을 받는다.**
    # 자기 기준에는 미달이 없기 때문이다. 그래서 같은 선종 배들의 최빈 주기를
    # 함께 재고, 자기 주기가 선종 표준에 못 미치는 배를 따로 표시한다.
    type_norm = {}
    for t in {v['g'] for v in vessels}:
        c = Counter(v['cadence'] for v in vessels if v['g'] == t and v['cadence'] > 0)
        type_norm[t] = c.most_common(1)[0][0] if c else 0
    for v in vessels:
        norm = type_norm.get(v['g'], 0)
        v['norm'] = norm
        v['belowNorm'] = bool(norm and v['cadence'] < norm)

    # ── 자기점검 ①: 국기별 척수 합 = 전체 ──────────────────────────────
    by_flag = Counter(v['f'] for v in vessels)
    if sum(by_flag.values()) != len(vessels):
        errors.append('국기별 합이 전체와 다르다')

    # ── 자기점검 ②: 선종별 척수 합 = 전체 ─────────────────────────────
    by_type = Counter(v['g'] for v in vessels)
    if sum(by_type.values()) != len(vessels):
        errors.append('선종별 합이 전체와 다르다')

    # ── 자기점검 ③: FFA 가 NOT REPORTING 이라 한 배는 결손이 있어야 한다 ──
    for v in vessels:
        if v['mtu'].startswith('NOT REPORTING') and v['short'] == 0 and v['zero'] == 0:
            # 항내 정박 등으로 정상 주기를 유지한 채 미보고 표기된 경우가 있다.
            # 오류로 세지 않고 사유를 남긴다.
            v['note'] = 'FFA 는 미보고로 표기했으나 일별 건수는 결손이 없다'

    # ── 한국 선단 ────────────────────────────────────────────────────
    kor = [v for v in vessels if v['f'] == '한국']
    kor_by_owner = defaultdict(list)
    for v in kor:
        if v['g'] == '선망':
            kor_by_owner[v['ownerKey']].append(v)
    owners = sorted(
        (
            {
                '선주': Counter(x['owner'] for x in vs).most_common(1)[0][0],
                '표기수': len({x['owner'] for x in vs}),
                '척수': len(vs),
                '선박': sorted(x['n'] for x in vs),
            }
            for vs in kor_by_owner.values()
        ),
        key=lambda d: -d['척수'],
    )

    # ── 어창 용량: 단위가 섞여 있어 단위별로 나눠 담는다 ──────────────────
    holds = []
    for v in kor:
        sp = specs.get(v['n'])
        if not sp:
            continue
        cap, unit = sp[G['Fish Hold Capacity']], sp[G['Fish Hold Capacity Unit']]
        if cap is None or unit is None:
            continue
        holds.append({'선명': v['n'], '선종': v['g'], '용량': float(cap),
                      '단위': '㎥' if 'Cubic' in str(unit) else 't'})
    hold_units = Counter(h['단위'] for h in holds)
    if len(hold_units) < 2:
        errors.append('어창 단위가 한 갈래다 — 원본 확인 필요(섞여 있어야 정상)')

    if errors:
        print('원본 대조 실패:', file=sys.stderr)
        for e in errors:
            print('  -', e, file=sys.stderr)
        return 1

    not_reporting = [v for v in vessels if v['mtu'].startswith('NOT REPORTING')]
    out = {
        '_meta': {
            '출처': 'FFA 어선등록부 주간 VMS 보고현황 보고서 (VRST)',
            '기간': f'{days[0]} ~ {days[-1]}',
            # 명부는 추출 시점 기준이라 기간 이후 등록분이 함께 실린다. 보고할 날이
            # 없었으므로 집계에서 뺐다 — 총척수가 명부 행수보다 그만큼 적은 이유다.
            '기간후등록제외': [f"{v['선박']}({v['등록']})" for v in after_period],
            '등급': 'A',
            '주의': (
                '원본 집계 시트(Countbyflag)의 선종 열 합이 실제 척수와 다르다 — '
                + ' · '.join(type_gap)
                + '. 이 페이지는 집계 시트가 아니라 선박별 원표에서 직접 세었다.'
            ),
            '측정경계': (
                '조업허가(Good Standing)는 FFA 회원국 수역 조업 자격이다. 실제 조업 여부나 '
                '어획량이 아니다. 일별 숫자는 VMS 위치보고 건수이며 조업일이 아니다.'
            ),
            '단위경고': '어창 용량은 ㎥ 와 t 가 섞여 있다. 두 단위를 더하거나 한 축에 얹지 않는다.',
            '갱신방법': 'python3 scripts/build_ffa_vrst.py <VRST xlsx>',
        },
        '요약': {
            '총척수': len(vessels),
            '국기수': len(by_flag),
            '미보고척수': len(not_reporting),
            '한국척수': len(kor),
            '일수': len(days),
        },
        '국기별': [
            {'국기': f, '척수': n, **{t: sum(1 for v in vessels if v['f'] == f and v['g'] == t)
                                     for t in sorted(by_type)}}
            for f, n in by_flag.most_common()
        ],
        '선종별': [{'선종': t, '척수': n} for t, n in by_type.most_common()],
        '선종표준': [{'선종': t, '표준주기': n} for t, n in sorted(type_norm.items(), key=lambda kv: -kv[1])],
        '표준미달': sorted(
            ({'선명': v['n'], '국기': v['f'], '선종': v['g'], '자체주기': v['cadence'],
              '선종표준': v['norm'], 'VMS': v['mtu'] or '정상'}
             for v in vessels if v['belowNorm']),
            key=lambda d: (d['자체주기'] / d['선종표준'], d['선명']),
        ),
        '미보고': sorted(
            ({'선명': v['n'], '국기': v['f'], '선종': v['g'], '사유': v['mtu'],
              '무보고일': v['zero'], '결손일': v['short'], '자체주기': v['cadence'],
              '선종표준': v['norm'], '표준미달': v['belowNorm'], '적용일': v['days적용']}
             for v in not_reporting),
            key=lambda d: (-d['결손일'], d['선명']),
        ),
        '한국선단': {
            '선종별': [{'선종': t, '척수': n} for t, n in Counter(v['g'] for v in kor).most_common()],
            '선주별': owners,
            '어창': sorted(holds, key=lambda h: (h['단위'], -h['용량'])),
            '어창단위': dict(hold_units),
        },
        '일자': days,
    }

    outdir = Path(args.out)
    (outdir / 'ffa_vrst_v1.json').write_text(
        json.dumps(out, ensure_ascii=False, indent=1), encoding='utf-8')

    # 탐색기용 등록부 (기존 FleetRegistryExplorer 스키마 그대로)
    rows = []
    for v in vessels:
        sp = specs.get(v['n'])
        loa = None
        if sp and sp[G['Vessel Length (LOA)']]:
            m = re.match(r'([\d.]+)', str(sp[G['Vessel Length (LOA)']]))
            loa = float(m.group(1)) if m else None
        rows.append({
            'o': 'FFA', 'n': v['n'], 'f': v['f'], 'g': v['g'],
            't': float(v['grt']) if isinstance(v['grt'], (int, float)) else None,
            'y': int(v['built']) if isinstance(v['built'], (int, float)) else None,
            'l': loa,
            'w': v['owner'] or None,
            'p': str(sp[G['Operator Name']]) if sp and sp[G['Operator Name']] else None,
            'h': str(sp[G['Port of Registry']]) if sp and sp[G['Port of Registry']] else None,
            'e': f"{v['st']} · VMS {v['mtu'] or '표기없음'}",
        })
    (outdir / 'ffa_fleet_db_v1.json').write_text(json.dumps({
        '_meta': {
            '생성일': days[-1],
            '출처': out['_meta']['출처'] + f" ({out['_meta']['기간']})",
            '등급': 'A',
            '키': {'o': '기구', 'n': '선명', 'f': '선적', 'g': '선종', 't': '총톤수',
                   'y': '건조년', 'l': '전장(m)', 'w': '소유사', 'p': '운영사',
                   'h': '등록항', 'e': '자격·VMS'},
            '행수': len(rows),
            '주의': out['_meta']['측정경계'],
            '기구별': {'FFA': len(rows)},
            '갱신방법': out['_meta']['갱신방법'],
        },
        'rows': rows,
    }, ensure_ascii=False), encoding='utf-8')

    print(f'선박 {len(vessels)}척 · 국기 {len(by_flag)} · 미보고 {len(not_reporting)}척 · '
          f'한국 {len(kor)}척 · 기간 {days[0]}~{days[-1]}')
    print('집계표 불일치(원표 채택):', ' · '.join(type_gap) or '없음')
    return 0


if __name__ == '__main__':
    raise SystemExit(main())
