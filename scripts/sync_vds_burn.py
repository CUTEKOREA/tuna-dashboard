#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""수역별 회사별 조업일수 소진현황 xlsx -> data/vds_company_burn.json.

원자료: 미경실 조업일수 대장 (연도 시트 2014~2026, 수역 블록별 원장 + 요약 4행).
- 각 수역 블록의 **마지막** 소진일수/잔여일수/소진률 세트와 그 위 가용일수 행을 요약으로 쓴다.
- Kiribati 는 세 블록(20-21어기 이월분 / 회사집계 / 집계) 중 "집계" 블록이 최종 -
  (20-21어기)·회사집계 블록은 수역 키가 달라 자동 제외된다.
- 소진률 셀은 비율(0.9=90%)이므로 x100. #DIV/0! 등 문자열은 가용·소진으로 재계산.
- 원문 소계는 그대로 두고 재계산과 1%p 넘게 어긋나면 경고만 출력한다 (임의 보정 금지).

사용: python3 scripts/sync_vds_burn.py "<xlsx 경로>" [--years 2023 2024 2025 2026]
"""
from __future__ import annotations

import argparse
import hashlib
import json
import sys
from pathlib import Path

import openpyxl

ZONES = {
    "PNG": "PNG",
    "Solomon": "Solomon",
    "Kiribati\n집계": "Kiribati",
    "Tuvalu": "Tuvalu",
    "Nauru": "Nauru",
    "FSM": "FSM",
}
SUMMARY_CONSUMED = {"소진일수"}
SUMMARY_REMAINING = {"잔여일수"}
SUMMARY_RATE = {"소진률(%)", "소진율(%)", "소진율"}
AVAILABLE_LABELS = {"총가용일수", "가용일수", "소계", "계", "총가용"}


def as_num(value):
    return round(float(value), 2) if isinstance(value, (int, float)) else None


def parse_sheet(ws):
    """수역 블록별 (회사명 리스트, 요약 dict) 추출."""
    companies: list[str] = []
    blocks: dict[str, dict] = {}
    zone_raw = None
    rows = list(ws.iter_rows(values_only=True))
    for r in rows:
        b = str(r[1]).strip() if r[1] is not None and str(r[1]).strip() else None
        label = str(r[2]).strip() if r[2] is not None else None
        if b == "구분":
            companies = []
            for cell in r[3:]:
                name = str(cell).strip() if cell else ""
                if name in ("계", "비고") or not name:
                    break
                companies.append(name)
            continue
        if b and not b.startswith("○"):
            zone_raw = b
        if zone_raw not in ZONES or not label:
            continue
        zone = ZONES[zone_raw]
        n = len(companies)
        vals = [as_num(v) for v in r[3:3 + n]]
        total = as_num(r[3 + n])
        entry = blocks.setdefault(zone, {})
        if label in AVAILABLE_LABELS:
            entry["available"] = (vals, total)
        elif label in SUMMARY_CONSUMED:
            entry["consumed"] = (vals, total)  # 마지막 세트가 남는다
        elif label in SUMMARY_REMAINING:
            entry["remaining"] = (vals, total)
        elif label in SUMMARY_RATE:
            entry["rate"] = (vals, total)
    return companies, blocks


ZONE_KO = {"PNG": "PNG", "Solomon": "솔로몬", "Kiribati": "키리바시",
           "Tuvalu": "투발루", "Nauru": "나우루", "FSM": "FSM"}


def parse_events(ws, year: int, companies_n: int) -> list[dict]:
    """원장 행(추가구매·전배)에서 최근 이벤트 추출 - 비고의 'M월 D일' 사용."""
    import re as _re
    events = []
    zone_raw = None
    for r in ws.iter_rows(values_only=True):
        b = str(r[1]).strip() if r[1] is not None and str(r[1]).strip() else None
        label = str(r[2]).strip() if r[2] is not None else None
        if b and not b.startswith("○") and b != "구분":
            zone_raw = b
        if zone_raw not in ZONES or not label:
            continue
        if "구매" not in label and "전배" not in label:
            continue
        note = str(r[3 + companies_n + 1] or "")
        m = _re.search(r"(\d{1,2})월\s*(\d{1,2})일", note)
        if not m:
            continue
        total = r[3 + companies_n]
        kind = "추가 구매" if "구매" in label else "전배"
        zone_ko = ZONE_KO[ZONES[zone_raw]]
        amount = f" ({abs(total):g}일)" if isinstance(total, (int, float)) and total else ""
        events.append({
            "date": f"{int(m.group(1)):02d}/{int(m.group(2)):02d}",
            "sortKey": int(m.group(1)) * 100 + int(m.group(2)),
            "msg": f"{zone_ko} 조업일수 {kind}{amount}",
        })
    events.sort(key=lambda e: -e["sortKey"])
    return [{k: e[k] for k in ("date", "msg")} for e in events[:5]]


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("source")
    parser.add_argument("--years", nargs="*", type=int, default=[2023, 2024, 2025, 2026])
    parser.add_argument("--asof", default="2026-08-28")
    parser.add_argument("--output", default="data/vds_company_burn.json")
    args = parser.parse_args()

    src = Path(args.source)
    wb = openpyxl.load_workbook(src, data_only=True)
    years_out = {}
    for year in args.years:
        ws = wb[str(year)]
        companies, blocks = parse_sheet(ws)
        zones_out = {}
        for zone, entry in blocks.items():
            if not {"consumed", "remaining"} <= set(entry):
                print(f"경고: {year} {zone} 요약 행 누락 - 건너뜀", file=sys.stderr)
                continue
            consumed, consumed_total = entry["consumed"]
            remaining, remaining_total = entry["remaining"]
            available = entry.get("available", (None, None))
            rate = entry.get("rate", (None, None))
            comp_out = {}
            for idx, name in enumerate(companies):
                c = consumed[idx] if idx < len(consumed) else None
                rem = remaining[idx] if idx < len(remaining) else None
                if c is None and rem is None:
                    continue
                avail = available[0][idx] if available[0] and idx < len(available[0]) else None
                if avail is None and c is not None and rem is not None:
                    avail = round(c + rem, 2)
                raw_rate = rate[0][idx] if rate[0] and idx < len(rate[0]) else None
                rate_pct = (round(raw_rate * 100) if raw_rate is not None
                            else (round(c / avail * 100) if avail else None))
                # 원문 소진률 대조 - 1%p 초과 어긋나면 경고 (보정하지 않음)
                if raw_rate is not None and avail:
                    recalc = c / avail * 100
                    if abs(recalc - raw_rate * 100) > 1:
                        print(f"경고: {year} {zone} {name} 소진률 원문 {raw_rate*100:.1f}% vs "
                              f"재계산 {recalc:.1f}%", file=sys.stderr)
                comp_out[name] = {
                    "available": avail,
                    "consumed": c,
                    "remaining": rem,
                    "ratePct": rate_pct,
                }
            zones_out[zone] = {
                "companies": comp_out,
                "total": {
                    "available": (available[1]
                                  if available[1] is not None
                                  else (round(consumed_total + remaining_total, 2)
                                        if consumed_total is not None and remaining_total is not None
                                        else None)),
                    "consumed": consumed_total,
                    "remaining": remaining_total,
                    "ratePct": round(rate[1] * 100) if rate[1] is not None else None,
                },
            }
        years_out[str(year)] = {"companies": companies, "zones": zones_out}

    latest_year = max(args.years)
    latest_ws = wb[str(latest_year)]
    latest_companies, _ = parse_sheet(latest_ws)
    recent_events = parse_events(latest_ws, latest_year, len(latest_companies))

    payload = {
        "asOf": args.asof,
        "recentEvents": recent_events,
        "source": {
            "file": src.name,
            "sha256": hashlib.sha256(src.read_bytes()).hexdigest(),
            "note": "미경실 조업일수 대장 - 수역 요약은 각 블록 최종 소진/잔여/소진률 행, Kiribati 는 집계 블록",
        },
        "years": years_out,
    }
    out = Path(args.output)
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(json.dumps(payload, ensure_ascii=False, indent=1) + "\n", encoding="utf-8")
    y = years_out.get("2026", {}).get("zones", {})
    print(f"OK {out} | 연도 {list(years_out)} | 2026 수역 {list(y)}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
