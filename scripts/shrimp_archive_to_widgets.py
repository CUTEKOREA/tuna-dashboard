#!/usr/bin/env python3
"""Build the shrimp dashboard v4 snapshot from the read-only Drive archive.

The converter intentionally has no network or agri_pipeline dependency.  It
reads the archived CSV/XLSX/Markdown/HTML sources, validates the six canonical
production assertions plus the JSON contract, and writes one atomic JSON file.
"""

from __future__ import annotations

import argparse
import copy
import csv
import html
import json
import os
import re
import sys
import tempfile
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Iterable, Mapping, Sequence

from openpyxl import load_workbook


ARCHIVE = Path(
    "/Users/idong-geon/Library/CloudStorage/GoogleDrive-cutekorea@gmail.com/"
    "내 드라이브/agri_data/01_수산물(Seafood)/shrimp/00_새우_관련자료"
)
REPO_ROOT = Path(__file__).resolve().parents[1]
V3_PATH = REPO_ROOT / "public/data/shrimp_real_data_v3.json"
OUTPUT_PATH = REPO_ROOT / "public/data/shrimp_real_data_v4.json"

FISHSTAT_PATH = (
    ARCHIVE
    / "10_원본데이터셋/FAO_FishStat/snapshot_2026-07-06/"
    "FishStat_2026.1.0_global_production_shrimp.csv"
)
SAGYP_PATH = (
    ARCHIVE
    / "01_자연산_어획·자원/Argentina_SAGyP/"
    "20260804-Argentina_Secretaría_de_Agricultura_Ganadería_y_Pesca-"
    "Argentina_Maritime_Landings_2026_through_04_August/"
    "260804_Desembarques_2026.xlsx"
)
AVANTI_PATH = (
    ARCHIVE
    / "04_가공·완제품·기업/Avanti_Feeds/"
    "20260724-Avanti_Feeds-Avanti_Feeds_Annual_Report_2025-26.md"
)
GLOBEFISH_PATH = (
    ARCHIVE
    / "05_시장·소매·소비/FAO_GLOBEFISH/"
    "20260525-FAO_GLOBEFISH-GLOBEFISH_Quarterly_Shrimp_Analysis_-_May_2026.md"
)
CNA_PATH = (
    ARCHIVE
    / "03_무역·가격/Ecuador_CNA/"
    "20260701-Cámara_Nacional_de_Acuacultura-"
    "CNA_Ecuador_Shrimp_Export_Statistics_-_May_2026.xlsx"
)
INFOFISH_PATH = (
    ARCHIVE
    / "09_이벤트·컨퍼런스/INFOFISH/"
    "20260306-INFOFISH-INFOFISH_International_Issue_2_2026.md"
)
KCS_PATH = (
    ARCHIVE
    / "03_무역·가격/Korea_KCS/snapshot_2026-07-06/kcs/"
    "KCS_2026YTD_HS_shrimp.csv"
)
NITEMTRADE_PATH = (
    ARCHIVE
    / "03_무역·가격/Korea_KCS/extracts/"
    "20260816-Korea_Customs-nitemtrade_shrimp_HS_partner_202401-202606.csv"
)
SERIES_ORIGIN_ORDER = ("에콰도르", "베트남", "인도", "태국", "중국")
SERIES_KCS_NAME = {
    "에콰도르": "에쿠아도르",  # nitemtrade 표기
    "베트남": "베트남",
    "인도": "인도",
    "태국": "태국",
    "중국": "중국",
}
SERIES_FAO_COUNTRIES = {
    "에콰도르": "Ecuador",
    "베트남": "Viet Nam",
    "인도": "India",
    "태국": "Thailand",
    "중국": "China",
    "한국": "Republic of Korea",
}
HS_MATRIX_PATH = (
    ARCHIVE
    / "03_무역·가격/Korea_KCS/snapshot_2026-07-06/hs/"
    "HS_matrix_shrimp.csv"
)
PINKSHEET_PATH = ARCHIVE / "03_무역·가격/legacy_raw_data/PinkSheet_Shrimp.csv"
PRICE_REPORT_PATH = (
    ARCHIVE
    / "03_무역·가격/FAO_GLOBEFISH_prices/"
    "20260717-FAO_GLOBEFISH-European_Fish_Price_Report_June_2026.md"
)
SOFIA_PATH = (
    ARCHIVE
    / "08_국가·산업리포트/FAO_SOFIA/"
    "20260617-FAO-The_State_of_World_Fisheries_and_Aquaculture_2026.md"
)

AQUACULTURE_SOURCES = frozenset({"BRACKISHWATER", "FRESHWATER", "MARINE"})
ALLOWED_TELEMETRY = frozenset({"SYNCED", "STATIC"})
ALLOWED_PILLARS = frozenset({"S1", "S2", "S3", "S4", "S5"})
EXPECTED_PILLARS = {"S1": 6, "S2": 4, "S3": 5, "S4": 6, "S5": 3}
EXPECTED_WIDGET_COUNT = sum(EXPECTED_PILLARS.values())
BANNED_STRINGS = (
    "illustrative",
    "자체 추정",
    "업계 추정",
    "자체 합성",
    "NotebookLM",
    "LIVE API 연동",
    "실시간 연동중",
)
UNFILTERED_VALUES = ("9501198", "3267045", "12768242")

COUNTRY_KO = {
    "Argentina": "아르헨티나",
    "Bangladesh": "방글라데시",
    "Brazil": "브라질",
    "China": "중국",
    "Ecuador": "에콰도르",
    "France": "프랑스",
    "India": "인도",
    "Indonesia": "인도네시아",
    "Japan": "일본",
    "Mexico": "멕시코",
    "Netherlands (Kingdom of the)": "네덜란드",
    "Republic of Korea": "한국",
    "Spain": "스페인",
    "Thailand": "태국",
    "United States of America": "미국",
    "Venezuela (Bolivarian Republic of)": "베네수엘라",
    "Viet Nam": "베트남",
}
SPECIES_KO = {
    "Whiteleg shrimp": "흰다리새우",
    "Giant tiger prawn": "블랙타이거",
    "Natantian decapods NEI": "새우류 기타",
    "Akiami paste shrimp": "젓새우",
    "Northern prawn": "북방새우",
}


class DataContractError(RuntimeError):
    """Raised when an archive source or generated payload violates the spec."""


@dataclass(frozen=True)
class GateResult:
    name: str
    actual: Decimal
    expected: Decimal
    tolerance: Decimal

    @property
    def passed(self) -> bool:
        return abs(self.actual - self.expected) <= self.tolerance


def _require_file(path: Path) -> Path:
    if not path.is_file():
        raise FileNotFoundError(f"required archive source not found: {path}")
    return path


def _read_text(path: Path) -> str:
    return _require_file(path).read_text(encoding="utf-8")


def _read_csv_dicts(path: Path) -> list[dict[str, str]]:
    with _require_file(path).open(encoding="utf-8-sig", newline="") as stream:
        return list(csv.DictReader(stream))


def _read_csv_rows(path: Path) -> list[list[str]]:
    with _require_file(path).open(encoding="utf-8-sig", newline="") as stream:
        return list(csv.reader(stream))


def _decimal(value: Any, *, context: str) -> Decimal:
    if value is None or str(value).strip() == "":
        raise DataContractError(f"missing numeric value: {context}")
    normalized = str(value).strip().replace(" ", "").replace(",", "")
    try:
        return Decimal(normalized)
    except InvalidOperation as exc:
        raise DataContractError(f"invalid numeric value for {context}: {value!r}") from exc


def _json_number(value: Decimal, digits: int = 3) -> int | float:
    rounded = round(value, digits)
    if rounded == rounded.to_integral_value():
        return int(rounded)
    return float(rounded)


def _pct(numerator: Decimal, denominator: Decimal, digits: int = 1) -> float:
    if denominator == 0:
        raise DataContractError("percentage denominator is zero")
    return round(float(numerator / denominator * Decimal(100)), digits)


def _normalize_space(value: str) -> str:
    return re.sub(r"\s+", " ", value).strip()


def _require_match(pattern: str, text: str, *, label: str, flags: int = 0) -> re.Match[str]:
    match = re.search(pattern, text, flags)
    if not match:
        raise DataContractError(f"could not parse {label}")
    return match


def _unique_glob(directory: Path, pattern: str) -> Path:
    matches = sorted(directory.glob(pattern))
    if len(matches) != 1:
        raise DataContractError(
            f"expected one file for {directory / pattern}, found {len(matches)}"
        )
    return matches[0]


def filter_shrimp_species(rows: Iterable[dict[str, str]]) -> list[dict[str, str]]:
    """Keep only ISSCAAP 'Shrimps, prawns'; freshwater crustaceans stay out."""
    filtered = [
        row
        for row in rows
        if row.get("SPECIES.ISSCAAP_Group_En") == "Shrimps, prawns"
    ]
    if not filtered:
        raise DataContractError("FishStat shrimp species filter returned zero rows")
    return filtered


def cna_resumen_monthly(ws: Any) -> list[dict[str, Any]]:
    """Read only RESUMEN rows 10-122 and columns AB-AE."""
    if ws.title != "RESUMEN":
        raise DataContractError(f"expected CNA RESUMEN sheet, got {ws.title!r}")
    records: list[dict[str, Any]] = []
    for row_number in range(10, 123):
        month, pounds, dollars, price = (
            ws.cell(row=row_number, column=column).value for column in range(28, 32)
        )
        if not isinstance(month, datetime):
            raise DataContractError(
                f"CNA RESUMEN AB{row_number} is not a monthly date: {month!r}"
            )
        records.append(
            {
                "month": month,
                "pounds": _decimal(pounds, context=f"CNA AC{row_number}"),
                "dollars": _decimal(dollars, context=f"CNA AD{row_number}"),
                "price": _decimal(price, context=f"CNA AE{row_number}"),
            }
        )
    if len(records) != 113:
        raise DataContractError(f"CNA monthly slice has {len(records)} rows, expected 113")
    return records


def truncate_pinksheet(rows: Iterable[dict[str, str]]) -> list[dict[str, str]]:
    """Keep the nominal Pink Sheet series through 2023M10, before 1079 corruption."""
    kept: list[dict[str, str]] = []
    for row in rows:
        period = row.get("Period", "")
        if not re.fullmatch(r"\d{4}M\d{2}", period):
            raise DataContractError(f"invalid Pink Sheet period: {period!r}")
        if period <= "2023M10":
            kept.append(row)
    if not kept or kept[-1]["Period"] != "2023M10":
        raise DataContractError("Pink Sheet cutoff 2023M10 was not found")
    return kept


def _load_fishstat() -> tuple[list[dict[str, str]], dict[str, Decimal]]:
    raw_rows = _read_csv_dicts(FISHSTAT_PATH)
    required = {
        "PERIOD",
        "VALUE",
        "PRODUCTION_SOURCE_DET.CODE",
        "SPECIES.Name_En",
        "SPECIES.ISSCAAP_Group_En",
        "COUNTRY.Name_En",
    }
    if not raw_rows or not required.issubset(raw_rows[0]):
        raise DataContractError("FishStat columns do not match the required schema")

    shrimp_rows = filter_shrimp_species(raw_rows)
    rows_2024 = [row for row in shrimp_rows if row["PERIOD"] == "2024"]
    if not rows_2024:
        raise DataContractError("FishStat has no filtered 2024 rows")

    aquaculture = sum(
        (
            _decimal(row["VALUE"], context="FishStat 2024 aquaculture")
            for row in rows_2024
            if row["PRODUCTION_SOURCE_DET.CODE"] in AQUACULTURE_SOURCES
        ),
        Decimal(0),
    )
    capture = sum(
        (
            _decimal(row["VALUE"], context="FishStat 2024 capture")
            for row in rows_2024
            if row["PRODUCTION_SOURCE_DET.CODE"] == "CAPTURE"
        ),
        Decimal(0),
    )
    total = aquaculture + capture
    whiteleg = sum(
        (
            _decimal(row["VALUE"], context="FishStat 2024 whiteleg")
            for row in rows_2024
            if row["SPECIES.Name_En"] == "Whiteleg shrimp"
        ),
        Decimal(0),
    )
    return shrimp_rows, {
        "aquaculture_2024": aquaculture,
        "capture_2024": capture,
        "global_production_2024": total,
        "whiteleg_2024": whiteleg,
        "whiteleg_share_2024": whiteleg / total * Decimal(100),
    }


def _parse_sofia_aquaculture() -> Decimal:
    text = _read_text(SOFIA_PATH)
    line = _require_match(
        r"^\s*Marine shrimps\s+(.+?)\s+61\.5\s*$",
        text,
        label="SOFIA Marine shrimps row",
        flags=re.MULTILINE,
    ).group(1)
    values = re.findall(r"\d{1,3}(?: \d{3})?", line)
    if len(values) < 9:
        raise DataContractError(f"SOFIA Marine shrimps row has too few values: {line!r}")
    return _decimal(values[-1], context="SOFIA latest Marine shrimps") * Decimal(1000)


def canonical_gate_results(metrics: Mapping[str, Decimal]) -> list[GateResult]:
    aquaculture = metrics["aquaculture_2024"]
    capture = metrics["capture_2024"]
    global_production = metrics["global_production_2024"]
    return [
        GateResult("aquaculture 2024", aquaculture, Decimal(8_810_922), Decimal(1)),
        GateResult("capture 2024", capture, Decimal(3_135_769), Decimal(1)),
        GateResult(
            "global production 2024", global_production, Decimal(11_946_690), Decimal(1)
        ),
        GateResult(
            "aquaculture + capture",
            aquaculture + capture,
            global_production,
            Decimal(1),
        ),
        GateResult(
            "whiteleg share",
            metrics["whiteleg_share_2024"],
            Decimal("64.1"),
            Decimal("0.1"),
        ),
        GateResult(
            "SOFIA cross-check",
            aquaculture,
            metrics["sofia_aquaculture"],
            Decimal(1000),
        ),
    ]


def _fishstat_widgets(
    rows: Sequence[dict[str, str]], metrics: Mapping[str, Decimal]
) -> tuple[dict[str, Any], dict[str, Any], dict[str, Any]]:
    by_year: dict[str, dict[str, list[Decimal]]] = defaultdict(
        lambda: {"aquaculture": [], "capture": []}
    )
    for row in rows:
        bucket = None
        if row["PRODUCTION_SOURCE_DET.CODE"] in AQUACULTURE_SOURCES:
            bucket = "aquaculture"
        elif row["PRODUCTION_SOURCE_DET.CODE"] == "CAPTURE":
            bucket = "capture"
        if bucket:
            by_year[row["PERIOD"]][bucket].append(
                _decimal(row["VALUE"], context=f"FishStat {row['PERIOD']} {bucket}")
            )

    production_data: list[dict[str, Any]] = []
    for year in range(1950, 2025):
        year_values = by_year.get(str(year), {"aquaculture": [], "capture": []})
        production_data.append(
            {
                "연도": str(year),
                "양식": (
                    _json_number(sum(year_values["aquaculture"], Decimal(0)))
                    if year_values["aquaculture"]
                    else None
                ),
                "자연산": (
                    _json_number(sum(year_values["capture"], Decimal(0)))
                    if year_values["capture"]
                    else None
                ),
            }
        )

    rows_2024 = [row for row in rows if row["PERIOD"] == "2024"]
    countries: dict[str, dict[str, Decimal]] = defaultdict(
        lambda: {"aquaculture": Decimal(0), "capture": Decimal(0)}
    )
    species: dict[str, Decimal] = defaultdict(Decimal)
    for row in rows_2024:
        value = _decimal(row["VALUE"], context="FishStat 2024 country/species")
        country = row["COUNTRY.Name_En"]
        if row["PRODUCTION_SOURCE_DET.CODE"] in AQUACULTURE_SOURCES:
            countries[country]["aquaculture"] += value
        elif row["PRODUCTION_SOURCE_DET.CODE"] == "CAPTURE":
            countries[country]["capture"] += value
        species[row["SPECIES.Name_En"]] += value

    top_countries = sorted(
        countries.items(), key=lambda item: item[1]["aquaculture"], reverse=True
    )[:10]
    country_data = [
        {
            "국가": COUNTRY_KO.get(country, country),
            "양식": _json_number(values["aquaculture"]),
            "자연산": _json_number(values["capture"]),
        }
        for country, values in top_countries
    ]

    top_species = sorted(species.items(), key=lambda item: item[1], reverse=True)[:5]
    species_total = sum(species.values(), Decimal(0))
    top_total = sum((value for _, value in top_species), Decimal(0))
    species_data = [
        {
            "name": SPECIES_KO.get(name, name),
            "value": _pct(value, species_total),
            "tonnes": _json_number(value),
        }
        for name, value in top_species
    ]
    species_data.append(
        {
            "name": "기타",
            "value": _pct(species_total - top_total, species_total),
            "tonnes": _json_number(species_total - top_total),
        }
    )

    whiteleg_share = round(float(metrics["whiteleg_share_2024"]), 1)
    total_rounded = round(metrics["global_production_2024"])
    aquaculture_share = round(
        float(metrics["aquaculture_2024"] / metrics["global_production_2024"] * 100), 1
    )
    common_source = (
        "FAO FishStat 2026.1.0, "
        "FishStat_2026.1.0_global_production_shrimp.csv"
    )
    w01 = {
        "id": "w01_paradigm_shift",
        "title": "양식과 자연산의 교차 (톤)",
        "subtitle": "FishStat 장기계열의 실측 연도만 집계했으며 결측값을 보간하지 않음",
        "chartType": "area",
        "pillar": "S1",
        "telemetry": "SYNCED",
        "syncDate": "FishStat 2026.1.0 (2024년 기준)",
        "source": common_source,
        "sourceQuote": (
            "SPECIES.ISSCAAP_Group_En=Shrimps, prawns; "
            "aquaculture=BRACKISHWATER+FRESHWATER+MARINE; capture=CAPTURE"
        ),
        "sit": (
            f"2024년 새우 총생산은 {total_rounded:,}톤이며 양식 비중은 "
            f"{aquaculture_share:.1f}%다. 1950~2024년 각 점은 FishStat 원시행의 "
            "연도별 합계이고 중간 연도를 생성하지 않았다."
        ),
        "strat": "공급 구조를 볼 때 양식과 자연산을 하나의 성장선으로 합치지 말고 서로 다른 생산 기반으로 읽어야 한다.",
        "xKey": "연도",
        "areas": [
            {"key": "양식", "name": "양식", "color": "#10b981"},
            {"key": "자연산", "name": "자연산", "color": "#38bdf8"},
        ],
        "data": production_data,
        "unit": "톤",
        "yUnit": "톤",
    }
    w02 = {
        "id": "w02_top10_by_source",
        "title": "상위 10개국 생산 — 양식과 자연산 (톤)",
        "subtitle": "2024년 양식 생산량 상위 10개국을 고정해 같은 국가의 자연산 어획을 병렬 비교",
        "chartType": "bar",
        "pillar": "S1",
        "telemetry": "SYNCED",
        "syncDate": "FishStat 2026.1.0 (2024년 기준)",
        "source": common_source,
        "sourceQuote": "PERIOD=2024; filtered ISSCAAP group; ranked by aquaculture only",
        "sit": (
            f"양식 1위 중국은 {int(round(top_countries[0][1]['aquaculture'])):,}톤이다. "
            "순위는 양식 생산량으로 고정했기 때문에 자연산 막대는 같은 국가의 공급 구조 차이를 보여준다. "
            "이 순위는 한국 창구 순위가 아니다."
        ),
        "strat": "생산 1위 원산지를 한국 주력 창구로 읽지 말고, 시리즈 6개국 역할·관세청 세번 위젯과 따로 대조한다.",
        "xKey": "국가",
        "bars": [
            {"key": "양식", "name": "양식", "color": "#10b981"},
            {"key": "자연산", "name": "자연산", "color": "#38bdf8"},
        ],
        "data": country_data,
        "unit": "톤",
        "yUnit": "톤",
    }
    w03 = {
        "id": "w03_species_concentration",
        "title": "종별 생산 집중도 (%)",
        "subtitle": "2024년 필터 후 생산량 기준 상위 5종과 잔차",
        "chartType": "pie",
        "pillar": "S1",
        "telemetry": "SYNCED",
        "syncDate": "FishStat 2026.1.0 (2024년 기준)",
        "source": common_source,
        "sourceQuote": (
            f"Whiteleg shrimp | 2024 | {_json_number(metrics['whiteleg_2024'])} tonnes"
        ),
        "sit": (
            f"흰다리새우가 2024년 필터 후 총생산의 {whiteleg_share:.1f}%를 차지한다. "
            "블랙타이거와 새우류 기타를 더해도 흰다리 단일 종의 규모에 미치지 못한다."
        ),
        "strat": "종별 집중도는 양식 확대의 규모와 함께 단일 종 질병·종묘 리스크가 전 세계 공급에 전이될 수 있음을 보여준다.",
        "xKey": "name",
        "data": species_data,
        "unit": "%",
        "yUnit": "%",
    }
    return w01, w02, w03


def _argentina_widget() -> dict[str, Any]:
    workbook = load_workbook(_require_file(SAGYP_PATH), read_only=True, data_only=True)
    try:
        month_ws = workbook["Especie_Mes"]
        fleet_ws = workbook["Especie_Flota"]
        month_row = next(
            (
                row
                for row in month_ws.iter_rows(min_row=1, max_col=10, values_only=True)
                if row[0] == "Langostino"
            ),
            None,
        )
        fleet_row = next(
            (
                row
                for row in fleet_ws.iter_rows(min_row=1, max_col=11, values_only=True)
                if row[0] == "Langostino"
            ),
            None,
        )
        total_row = next(
            (
                row
                for row in month_ws.iter_rows(min_row=1, max_col=10, values_only=True)
                if row[0] == "Total"
            ),
            None,
        )
        if month_row is None or fleet_row is None or total_row is None:
            raise DataContractError("SAGyP Langostino or Total row not found")
        headers = [month_ws.cell(row=6, column=column).value for column in range(2, 10)]
        month_values = [
            _decimal(value, context=f"SAGyP Langostino {headers[index]}")
            for index, value in enumerate(month_row[1:9])
        ]
        cumulative = _decimal(month_row[9], context="SAGyP Langostino cumulative")
        if abs(sum(month_values, Decimal(0)) - cumulative) > Decimal("0.001"):
            raise DataContractError("SAGyP monthly Langostino sum does not match total")
        fleet_headers = [fleet_ws.cell(row=7, column=column).value for column in range(2, 11)]
        fleet_values = {
            str(name): _decimal(value, context=f"SAGyP fleet {name}")
            for name, value in zip(fleet_headers, fleet_row[1:10])
        }
        all_landings = _decimal(total_row[9], context="SAGyP all landings total")
    finally:
        workbook.close()

    data = []
    for index, (month, value) in enumerate(zip(headers, month_values)):
        label = f"{index + 1}월"
        record: dict[str, Any] = {"월": label, "양륙량": _json_number(value)}
        if index == 7:
            record["월"] = "8월(1~4일)"
            record["partial"] = True
        data.append(record)

    tangoneros = fleet_values["Tangoneros"]
    altura = fleet_values["Altura"]
    costeros = fleet_values["Costeros"]
    rada = fleet_values["Rada o Ría"]
    return {
        "id": "w04_argentina_landings",
        "title": "아르헨티나 붉은새우 월별 양륙 (톤)",
        "subtitle": "2026년 월별 양륙이며 8월은 8/1~8/4 나흘치",
        "chartType": "bar",
        "pillar": "S1",
        "telemetry": "SYNCED",
        "syncDate": "SAGyP 2026-08-04",
        "source": "Argentina SAGyP, 260804_Desembarques_2026.xlsx",
        "sourceQuote": (
            "Especie_Mes | Langostino | "
            + " / ".join(f"{header}={float(value):.3f}" for header, value in zip(headers, month_values))
            + f" | Total={float(cumulative):.3f}"
        ),
        "sit": (
            f"1~8월 붉은새우 양륙은 {float(cumulative):,.3f}톤이며 8월 값은 나흘치다. "
            f"선단 구성은 Tangoneros {float(tangoneros):,.3f}톤({_pct(tangoneros, cumulative)}%), "
            f"Fresq. Altura {float(altura):,.3f}톤({_pct(altura, cumulative)}%), "
            f"Fresq. Costeros {float(costeros):,.3f}톤({_pct(costeros, cumulative)}%), "
            f"Rada o Ría {float(rada):,.3f}톤({_pct(rada, cumulative)}%)이다. "
            f"전체 해양 양륙 {float(all_landings):,.3f}톤 중 붉은새우는 {_pct(cumulative, all_landings)}%다."
        ),
        "strat": "부분월을 완결월과 같은 추세로 해석하지 않고, 자연산 공급의 계절 집중과 선단 구성을 함께 읽어야 한다.",
        "xKey": "월",
        "bars": [{"key": "양륙량", "name": "양륙량", "color": "#f59e0b"}],
        "data": data,
        "unit": "톤",
        "yUnit": "톤",
    }


def _processing_reversal_widget() -> dict[str, Any]:
    directory = ARCHIVE / "04_가공·완제품·기업/legacy_raw_data"
    path = _unique_glob(directory, "*1976-2023.csv")
    rows = _read_csv_rows(path)
    if not rows:
        raise DataContractError("processed production CSV is empty")
    header = rows[0]
    required = {
        "Country (Name)",
        "Commodity (Name)",
        "Element (Name)",
        "Unit (Name)",
        "[2013]",
        "[2023]",
    }
    if not required.issubset(header):
        raise DataContractError("processed production columns do not match the spec")
    country_index = header.index("Country (Name)")
    commodity_index = header.index("Commodity (Name)")
    year_2013_index = header.index("[2013]")
    year_2023_index = header.index("[2023]")
    totals: dict[str, dict[str, Decimal]] = defaultdict(
        lambda: {"2013": Decimal(0), "2023": Decimal(0)}
    )
    commodity_names: set[str] = set()
    included_rows = 0
    total_rows = 0
    citation_rows = 0
    for row in rows[1:]:
        if len(row) == 1:
            citation_rows += 1
            continue
        if len(row) != len(header):
            raise DataContractError(f"unexpected processed production row length: {len(row)}")
        country = row[country_index]
        if country.startswith("Totals - Tonnes"):
            total_rows += 1
            continue
        if row[header.index("Element (Name)")] != "Processed production":
            raise DataContractError(f"unexpected processed production element for {country}")
        commodity_names.add(row[commodity_index])
        totals[country]["2013"] += _decimal(
            row[year_2013_index] or "0", context=f"{country} processed 2013"
        )
        totals[country]["2023"] += _decimal(
            row[year_2023_index] or "0", context=f"{country} processed 2023"
        )
        included_rows += 1
    if total_rows != 1 or citation_rows != 1 or len(commodity_names) != 24:
        raise DataContractError(
            "processed production exclusion/schema mismatch: "
            f"totals={total_rows}, citations={citation_rows}, products={len(commodity_names)}"
        )

    # 두 해 각각의 상위 8개국 합집합. 2023년만 기준으로 자르면 최대 하락국(태국 -53%)이
    # 9위로 밀려 차트에서 사라지는데, SIT은 그 값을 인용한다 — 차트-텍스트 비동기가 된다.
    rank_2013 = sorted(totals.items(), key=lambda item: item[1]["2013"], reverse=True)[:8]
    rank_2023 = sorted(totals.items(), key=lambda item: item[1]["2023"], reverse=True)[:8]
    selected = {country for country, _ in rank_2013} | {country for country, _ in rank_2023}
    top = sorted(
        ((c, totals[c]) for c in selected), key=lambda item: item[1]["2023"], reverse=True
    )
    data = [
        {
            "국가": COUNTRY_KO.get(country, country),
            "2013": _json_number(values["2013"]),
            "2023": _json_number(values["2023"]),
        }
        for country, values in top
    ]

    def change(country: str) -> float:
        start = totals[country]["2013"]
        end = totals[country]["2023"]
        return round(float((end / start - 1) * Decimal(100)), 1)

    ecuador_ratio = totals["Ecuador"]["2023"] / totals["Ecuador"]["2013"]
    return {
        "id": "w08_processing_reversal",
        "title": "국가별 가공생산 역전 (톤)",
        "subtitle": "국가별 제품형태 행을 합산한 2013년과 2023년 비교 · 두 해 각각의 상위 8개국 합집합",
        "chartType": "bar",
        "pillar": "S2",
        "telemetry": "SYNCED",
        "syncDate": "FAO 가공생산 2023년",
        "source": f"FAO FishStat processed production, {path.name}",
        "sourceQuote": (
            f"174 country-product rows; {len(commodity_names)} distinct product labels; "
            "Totals - Tonnes – net product weight excluded"
        ),
        "sit": (
            f"2013~2023년 가공생산은 에콰도르가 {float(ecuador_ratio):.1f}배로 늘고 "
            f"인도·베트남도 증가했다. 같은 원자료 집계에서 태국은 {change('Thailand'):.1f}%, "
            f"중국은 {change('China'):.1f}% 감소했다. 지정된 GLOBEFISH·INFOFISH 문서에는 "
            "EMS·AHPND와 이 감소를 직접 연결하는 문장이 없어 차트는 수치 변화만 진술한다."
        ),
        "strat": "가공 지도의 이동은 생산량 변화로 먼저 확인하고, 질병 인과는 별도의 직접 근거가 확보될 때만 결합해야 한다.",
        "xKey": "국가",
        "bars": [
            {"key": "2013", "name": "2013년", "color": "#64748b"},
            {"key": "2023", "name": "2023년", "color": "#14b8a6"},
        ],
        "data": data,
        "unit": "톤",
        "yUnit": "톤",
    }


def _avanti_widget() -> dict[str, Any]:
    text = _read_text(AVANTI_PATH)
    normalized = _normalize_space(text)
    margins = re.findall(
        r"Segment mar(?:gin|- gin) improved from\s+([\d.]+)%\s+to\s+([\d.]+)%\.",
        normalized,
    )
    if len(margins) < 2:
        raise DataContractError(f"expected two Avanti segment margin sentences, found {len(margins)}")
    feed_old, feed_new = map(Decimal, margins[0])
    processing_old, processing_new = map(Decimal, margins[1])
    share = _require_match(
        r"estimated market share remains approximately\s+(\d+)%\s+to\s+(\d+)%",
        normalized,
        label="Avanti market share",
        flags=re.IGNORECASE,
    )
    share_low, share_high = map(int, share.groups())
    return {
        "id": "w09_feed_vs_processing_margin",
        "title": "사료와 가공의 마진 분리 (%)",
        "subtitle": (
            f"인도 새우사료 시장 {share_low}~{share_high}% 점유 1개사 실적 — 세계 대표값 아님"
        ),
        "chartType": "bar",
        "pillar": "S2",
        "telemetry": "STATIC",
        "syncDate": "Avanti Feeds FY2025-26",
        "source": "Avanti Feeds Annual Report 2025-26",
        "sourceQuote": (
            f"Segment margin improved from {feed_old}% to {feed_new}%. "
            f"Segment margin improved from {processing_old}% to {processing_new}%."
        ),
        "sit": (
            f"Avanti의 새우사료 부문 마진은 {feed_old}%에서 {feed_new}%로, 가공새우는 "
            f"{processing_old}%에서 {processing_new}%로 상승했다. 두 사업의 마진 수준과 개선폭이 다르다."
        ),
        "strat": "단일 기업 실적을 세계 산업 평균으로 확대하지 말고, 사료와 가공의 수익 구조가 분리돼 움직인다는 사례로 읽어야 한다.",
        "xKey": "사업부",
        "bars": [
            {"key": "FY2024-25", "name": "FY2024-25", "color": "#64748b"},
            {"key": "FY2025-26", "name": "FY2025-26", "color": "#10b981"},
        ],
        "data": [
            {
                "사업부": "새우사료",
                "FY2024-25": _json_number(feed_old, 2),
                "FY2025-26": _json_number(feed_new, 2),
            },
            {
                "사업부": "가공새우",
                "FY2024-25": _json_number(processing_old, 2),
                "FY2025-26": _json_number(processing_new, 2),
            },
        ],
        "unit": "%",
        "yUnit": "%",
    }


def _parse_exporters(text: str) -> list[dict[str, Any]]:
    block = text.split("World top exporters of shrimp", 1)[1].split("Source:", 1)[0]
    row_pattern = re.compile(
        r"^\s*(Ecuador|India|Viet Nam|Indonesia|China|Thailand|Argentina|Other countries|Total exports)"
        r"\s+([\d ]+\.\d{2})\s+([\d ]+\.\d{2})\s+([\d ]+\.\d{2})\s+([+-]?\d+)\s*$",
        re.MULTILINE,
    )
    rows = []
    for match in row_pattern.finditer(block):
        country, y2023, y2024, y2025, change = match.groups()
        rows.append(
            {
                "country": country,
                "2023": _decimal(y2023, context=f"exporters {country} 2023"),
                "2024": _decimal(y2024, context=f"exporters {country} 2024"),
                "2025": _decimal(y2025, context=f"exporters {country} 2025"),
                "change": int(change),
                "quote": _normalize_space(match.group(0)),
            }
        )
    if len(rows) != 9:
        raise DataContractError(f"GLOBEFISH exporter table parsed {len(rows)} rows, expected 9")
    return rows


def _parse_importers(text: str) -> list[dict[str, Any]]:
    block = text.split("World top importers of shrimp", 1)[1].split("Source:", 1)[0]
    number = r"([\d ]+\.\d{2})"
    patterns = {
        "China": rf"China\s+{number}\s+{number}\s+{number}\s+([+-]?\d+)",
        "United States of America": rf"United States of\s+{number}\s+{number}\s+{number}\s+([+-]?\d+)\s+America",
        "Japan": rf"Japan\s+{number}\s+{number}\s+{number}\s+([+-]?\d+)",
        "Spain": rf"Spain\s+{number}\s+{number}\s+{number}\s+([+-]?\d+)",
        "France": rf"France\s+{number}\s+{number}\s+{number}\s+([+-]?\d+)",
        "Republic of Korea": rf"Republic of\s+{number}\s+{number}\s+{number}\s+([+-]?\d+)\s+Korea",
        "Netherlands (Kingdom of the)": rf"Netherlands\s+{number}\s+{number}\s+{number}\s+([+-]?\d+)\s+\(Kingdom of the\)",
        "Other countries": rf"Other countries\s+{number}\s+{number}\s+{number}\s+([+-]?\d+)",
        "Total imports": rf"Total imports\s+{number}\s+{number}\s+{number}\s+([+-]?\d+)",
    }
    rows = []
    for country, pattern in patterns.items():
        match = _require_match(pattern, block, label=f"importer row {country}")
        y2023, y2024, y2025, change = match.groups()
        rows.append(
            {
                "country": country,
                "2023": _decimal(y2023, context=f"importers {country} 2023"),
                "2024": _decimal(y2024, context=f"importers {country} 2024"),
                "2025": _decimal(y2025, context=f"importers {country} 2025"),
                "change": int(change),
                "quote": _normalize_space(match.group(0)),
            }
        )
    return rows


def _globefish_widgets() -> tuple[dict[str, Any], dict[str, Any], dict[str, Any]]:
    text = _read_text(GLOBEFISH_PATH)
    exporters = _parse_exporters(text)
    importers = _parse_importers(text)
    export_chart_rows = [
        row for row in exporters if row["country"] not in {"Other countries", "Total exports"}
    ]
    import_chart_rows = [
        row for row in importers if row["country"] not in {"Other countries", "Total imports"}
    ]
    global_value = _require_match(
        r"Global imports of shrimp in 2025 totalled\s+[\d.]+\s+million tonnes valued at USD\s+([\d.]+)\s+billion",
        _normalize_space(text),
        label="GLOBEFISH global import value",
        flags=re.IGNORECASE,
    )
    us_detail = _require_match(
        r"Shrimp imports in the United States totalled\s+([\d ]+)\s+tonnes worth USD\s+([\d ]+)\s+million during 2025",
        _normalize_space(text),
        label="GLOBEFISH United States detail",
        flags=re.IGNORECASE,
    )
    data_export = [
        {
            "국가": COUNTRY_KO[row["country"]],
            "2023": _json_number(row["2023"], 2),
            "2024": _json_number(row["2024"], 2),
            "2025": _json_number(row["2025"], 2),
            "증감률": row["change"],
        }
        for row in export_chart_rows
    ]
    data_import = [
        {
            "시장": COUNTRY_KO[row["country"]],
            "2023": _json_number(row["2023"], 2),
            "2024": _json_number(row["2024"], 2),
            "2025": _json_number(row["2025"], 2),
            "증감률": row["change"],
        }
        for row in import_chart_rows
    ]
    common_bars = [
        {"key": "2023", "name": "2023년", "color": "#64748b"},
        {"key": "2024", "name": "2024년", "color": "#38bdf8"},
        {"key": "2025", "name": "2025년", "color": "#10b981"},
    ]
    w10 = {
        "id": "w10_world_exporters",
        "title": "세계 수출국 3개년 물량 (천 톤)",
        "subtitle": "FAO GLOBEFISH 표 원문 · 2023~2025년",
        "chartType": "bar",
        "pillar": "S3",
        "telemetry": "STATIC",
        "syncDate": "FAO GLOBEFISH 2025년",
        "source": "FAO GLOBEFISH Quarterly Shrimp Analysis, May 2026",
        "sourceQuote": " | ".join(row["quote"] for row in exporters),
        "sit": (
            f"2025년 에콰도르 수출은 {float(export_chart_rows[0]['2025']):,.2f}천 톤으로 전년보다 "
            f"{export_chart_rows[0]['change']:+d}% 늘었다. 아르헨티나는 "
            f"{next(row for row in exporters if row['country']=='Argentina')['change']:+d}%, "
            f"태국은 {next(row for row in exporters if row['country']=='Thailand')['change']:+d}%로 감소했다. "
            "세계 수출 순위는 한국 창구 순위가 아니다."
        ),
        "strat": "수출 1위 원산지를 한국 주력 창구로 읽지 말고, 관세청 세번·단가 위젯과 따로 대조한다.",
        "xKey": "국가",
        "bars": copy.deepcopy(common_bars),
        "data": data_export,
        "unit": "천 톤",
        "yUnit": "천 톤",
    }
    w14 = {
        "id": "w14_top_import_markets",
        "title": "3대 수입시장 규모 (천 톤)",
        "subtitle": "FAO GLOBEFISH 표 원문 · 7개 주요 시장의 2023~2025년 물량",
        "chartType": "bar",
        "pillar": "S4",
        "telemetry": "STATIC",
        "syncDate": "FAO GLOBEFISH 2025년",
        "source": "FAO GLOBEFISH Quarterly Shrimp Analysis, May 2026",
        "sourceQuote": " | ".join(row["quote"] for row in importers),
        "sit": (
            f"2025년 중국은 {float(import_chart_rows[0]['2025']):,.2f}천 톤으로 1% 감소했고 미국은 "
            f"{float(import_chart_rows[1]['2025']):,.2f}천 톤으로 2% 증가했다. 한국은 "
            f"{float(next(row for row in importers if row['country']=='Republic of Korea')['2025']):,.2f}천 톤이다. "
            f"미국 상세는 {_decimal(us_detail.group(1), context='US tonnes'):,}톤·"
            f"USD {_decimal(us_detail.group(2), context='US value'):,}M이다."
        ),
        "strat": "관세율을 이 표에서 역산하지 않고, 수입시장 물량 변화와 별도 정책 근거를 분리해 읽어야 한다.",
        "xKey": "시장",
        "bars": copy.deepcopy(common_bars),
        "data": data_import,
        "unit": "천 톤",
        "yUnit": "천 톤",
    }
    facts = {
        "global_trade_value_billion": _decimal(
            global_value.group(1), context="global trade value"
        ),
        "ecuador_2025_thousand_tonnes": export_chart_rows[0]["2025"],
        "ecuador_change": Decimal(export_chart_rows[0]["change"]),
    }
    return w10, w14, facts


def _cna_widget() -> tuple[dict[str, Any], dict[str, Decimal]]:
    workbook = load_workbook(_require_file(CNA_PATH), read_only=True, data_only=True)
    try:
        monthly = cna_resumen_monthly(workbook["RESUMEN"])
        market_ws = workbook["MERCADO PAÍS ACUM"]
        # 권역 비교표는 A열이 아니라 J열(라벨)·K/L열(2025/2026 물량 비중)이다.
        # A열 EUROPA 행의 K/L은 같은 줄의 RESTO DE ASIA 비중이 붙어 있어
        # 유럽을 4.7%로 오독한다.
        market_rows: dict[str, tuple[Decimal, Decimal]] = {}
        korea_share: Decimal | None = None
        for row_number in range(12, 80):
            bloc = market_ws.cell(row=row_number, column=10).value
            if bloc in {"CHINA", "EEUU", "EUROPA"} and bloc not in market_rows:
                market_rows[str(bloc)] = (
                    _decimal(market_ws.cell(row=row_number, column=11).value, context=f"CNA K{row_number}"),
                    _decimal(market_ws.cell(row=row_number, column=12).value, context=f"CNA L{row_number}"),
                )
            country = market_ws.cell(row=row_number, column=1).value
            if country == "COREA DEL SUR":
                korea_share = _decimal(
                    market_ws.cell(row=row_number, column=8).value,
                    context=f"CNA H{row_number} COREA DEL SUR",
                )
    finally:
        workbook.close()
    if set(market_rows) != {"CHINA", "EEUU", "EUROPA"}:
        raise DataContractError(f"CNA market rows incomplete: {sorted(market_rows)}")
    if korea_share is None:
        raise DataContractError("CNA COREA DEL SUR share was not found")
    first = monthly[0]
    last = monthly[-1]
    if first["month"].strftime("%Y-%m") != "2017-01" or last["month"].strftime("%Y-%m") != "2026-05":
        raise DataContractError("CNA monthly date range is not 2017-01 through 2026-05")
    europe_2026 = market_rows["EUROPA"][1]
    if not (Decimal("0.17") <= europe_2026 <= Decimal("0.19")):
        raise DataContractError(f"CNA Europe 2026 share out of range: {europe_2026}")
    data = [
        {
            "기간": row["month"].strftime("%Y-%m"),
            "수출량": _json_number(row["pounds"] / Decimal(1_000_000), 3),
            "단가": _json_number(row["price"], 4),
        }
        for row in monthly
    ]
    china_2026 = float(market_rows["CHINA"][1] * 100)
    us_2026 = float(market_rows["EEUU"][1] * 100)
    europe_2025_pct = float(market_rows["EUROPA"][0] * 100)
    europe_2026_pct = float(europe_2026 * 100)
    korea_pct = float(korea_share * 100)
    return (
        {
            "id": "w11_ecuador_monthly",
            "title": "에콰도르 월별 수출량과 단가 (백만 파운드 · 달러/파운드)",
            "subtitle": "CNA RESUMEN 10~122행만 사용한 2017년 1월~2026년 5월 113개월",
            "chartType": "composed",
            "pillar": "S3",
            "telemetry": "SYNCED",
            "syncDate": "CNA 2026년 5월",
            "source": "Cámara Nacional de Acuacultura, Export Statistics May 2026",
            "sourceQuote": (
                f"RESUMEN!AB122:AE122 | {last['month'].date()} | {last['pounds']} | "
                f"{last['dollars']} | {last['price']} | "
                f"COREA DEL SUR Part. Libras={korea_share}"
            ),
            "sit": (
                f"2026년 1~5월 목적지 비중은 중국 {china_2026:.1f}%, 미국 {us_2026:.1f}%, "
                f"유럽 {europe_2026_pct:.1f}%다. 유럽은 2025년 같은 누계 {europe_2025_pct:.1f}%에서 "
                f"낮아졌다. 한국(COREA DEL SUR)은 같은 표에서 {korea_pct:.2f}%다."
            ),
            "strat": (
                "한국 창구는 에콰도르 수출의 잔여 규격이다. 수출량과 파운드당 단가를 "
                "한 축으로 합치지 말고, 중국·미국 창구가 변할 때 한국으로 흘러오는 "
                "규격을 따로 읽는다. SECA 특혜는 발효·양허 확인 전 견적에 넣지 않는다."
            ),
            "xKey": "기간",
            "bars": [{"key": "수출량", "name": "수출량 (백만 파운드)", "color": "#10b981"}],
            "lines": [{"key": "단가", "name": "단가 (달러/파운드)", "color": "#f59e0b"}],
            "data": data,
            "unit": "백만 파운드 · 달러/파운드",
            "yUnit": "백만 파운드 · 달러/파운드",
        },
        {"korea_share": korea_share},
    )


def _infofish_widget() -> dict[str, Any]:
    source_lines = _read_text(INFOFISH_PATH).splitlines()
    start = next(
        (
            index
            for index, line in enumerate(source_lines)
            if "frozen shrimp imports into Vietnam" in line
        ),
        None,
    )
    if start is None:
        raise DataContractError("INFOFISH Vietnam paragraph was not found")
    # The archived PDF extraction preserves two page columns.  The relevant
    # paragraph occupies the right column, which begins at character 86.
    normalized = _normalize_space(
        " ".join(line[86:] for line in source_lines[start : start + 6] if len(line) > 86)
    )
    vietnam = _require_match(
        r"In 2025, frozen shrimp imports into Vietnam increased sharply, rising\s+(\d+)%\s+compared with 2024 to\s+([\d ]+)\s+tonnes\.\s+The main suppliers were Ecuador and India\.",
        normalized,
        label="INFOFISH Vietnam imports",
    )
    thailand = _require_match(
        r"Frozen shrimp imports into Thailand increased by\s+(\d+)%\s+in 2025 to\s+([\d ]+)\s+tonnes, with supplies sourced mainly from Ecuador and India\.",
        normalized,
        label="INFOFISH Thailand imports",
    )
    vn_change, vn_tonnes = int(vietnam.group(1)), _decimal(vietnam.group(2), context="Vietnam imports")
    th_change, th_tonnes = int(thailand.group(1)), _decimal(thailand.group(2), context="Thailand imports")
    return {
        "id": "w12_reprocessing_hubs",
        "title": "재가공 허브의 원료 수입 (톤)",
        "subtitle": "2025년 냉동 새우 원료 수입 · 공급원은 에콰도르와 인도",
        "chartType": "bar",
        "pillar": "S3",
        "telemetry": "STATIC",
        "syncDate": "INFOFISH 2025년",
        "source": "INFOFISH International Issue 2, 2026",
        "sourceQuote": f"{_normalize_space(vietnam.group(0))} {_normalize_space(thailand.group(0))}",
        "sit": (
            f"2025년 냉동 새우 수입은 베트남 {int(vn_tonnes):,}톤({vn_change:+d}%), "
            f"태국 {int(th_tonnes):,}톤({th_change:+d}%)이다. 두 나라 모두 주요 공급원은 에콰도르와 인도다. "
            "태국은 원물을 들여 가공하는 허브이고, 한국이 받는 태국산은 원물보다 조제품이 더 무겁다."
        ),
        "strat": "산지 완제품과 제3국 재가공을 한 공급자로 묶지 말고, 한국 창구는 세번(030617·160521)으로 나눈다.",
        "xKey": "허브",
        "bars": [{"key": "수입량", "name": "수입량", "color": "#0d9488"}],
        "data": [
            {"허브": "베트남", "수입량": int(vn_tonnes), "증감률": vn_change, "공급원": "에콰도르·인도"},
            {"허브": "태국", "수입량": int(th_tonnes), "증감률": th_change, "공급원": "에콰도르·인도"},
        ],
        "unit": "톤",
        "yUnit": "톤",
        "note": "원문에서 인도네시아와 중국에 동일하게 인쇄된 180,535톤은 차트에서 제외함",
    }


def _kcs_widget() -> tuple[dict[str, Any], dict[str, Decimal]]:
    matrix_rows = _read_csv_dicts(HS_MATRIX_PATH)
    stages = {row["hs6"]: row["stage"] for row in matrix_rows}
    rows = _read_csv_dicts(KCS_PATH)
    if len(rows) != 356:
        raise DataContractError(f"KCS has {len(rows)} rows, expected 356")
    detail = [row for row in rows if row["year"] != "총계"]
    used_hs = {row["hs_query"] for row in detail}
    expected_hs = {"030616", "030617", "160521", "160529"}
    if used_hs != expected_hs:
        raise DataContractError(f"KCS HS coverage mismatch: {sorted(used_hs)}")
    if any(stages.get(code) not in {"1_frozen", "3_prepared"} for code in used_hs):
        raise DataContractError("HS matrix lacks a frozen/prepared stage for a used KCS code")
    countries: dict[str, dict[str, Decimal]] = defaultdict(
        lambda: {"1_frozen": Decimal(0), "3_prepared": Decimal(0), "weight": Decimal(0)}
    )
    total_dollars = Decimal(0)
    total_weight = Decimal(0)
    for row in detail:
        dollars = _decimal(row["impDlr"], context="KCS impDlr")
        weight = _decimal(row["impWgt"], context="KCS impWgt")
        stage = stages[row["hs_query"]]
        country = row["statCdCntnKor1"]
        countries[country][stage] += dollars
        countries[country]["weight"] += weight
        total_dollars += dollars
        total_weight += weight
    top = sorted(
        countries.items(),
        key=lambda item: item[1]["1_frozen"] + item[1]["3_prepared"],
        reverse=True,
    )[:8]
    data = [
        {
            "교역국": country,
            "냉동": _json_number(values["1_frozen"] / Decimal(1_000_000), 3),
            "조제": _json_number(values["3_prepared"] / Decimal(1_000_000), 3),
        }
        for country, values in top
    ]
    return (
        {
            "id": "w13_kr_import_by_stage",
            "title": "한국 수입 — 냉동과 조제 (백만 달러)",
            "subtitle": "2026년 1~5월 누계 — 연환산·전년동기비 산출 불가",
            "chartType": "bar",
            "pillar": "S3",
            "telemetry": "SYNCED",
            "syncDate": "관세청 2026년 1~5월 누계",
            "source": "KCS 2026YTD HS shrimp + HS_matrix_shrimp.csv",
            "sourceQuote": (
                f"detail rows={len(detail)}; hs_query={','.join(sorted(used_hs))}; "
                f"impDlr={int(total_dollars)}; impWgt={int(total_weight)}"
            ),
            "sit": (
                f"4개 HS 상세행 합계는 USD {int(total_dollars):,}, {int(total_weight):,}kg이다. "
                f"베트남은 냉동 USD {int(countries['베트남']['1_frozen']):,}와 조제 USD "
                f"{int(countries['베트남']['3_prepared']):,}를 합쳐 최대 공급국이다. "
                "이 표는 2026년 1~5월 금액이다. 시리즈 6개국 물량·단가는 상반기 030617·160521 위젯에서 따로 본다."
            ),
            "strat": "0306 냉동과 1605 조제를 함께 보아야 재가공국의 역할과 한국 수입 구성을 동시에 설명할 수 있다.",
            "xKey": "교역국",
            "bars": [
                {"key": "냉동", "name": "0306 냉동", "color": "#38bdf8"},
                {"key": "조제", "name": "1605 조제", "color": "#f59e0b"},
            ],
            "data": data,
            "unit": "백만 달러",
            "yUnit": "백만 달러",
        },
        {"total_dollars": total_dollars, "total_weight": total_weight},
    )


def _pinksheet_widget() -> dict[str, Any]:
    raw_rows = _read_csv_dicts(PINKSHEET_PATH)
    rows = truncate_pinksheet(raw_rows)
    if len(rows) != 766:
        raise DataContractError(f"Pink Sheet has {len(rows)} valid rows, expected 766")
    first, last = rows[0], rows[-1]
    if (first["Period"], first["Price"]) != ("1960M01", "1.433003"):
        raise DataContractError(f"unexpected Pink Sheet first row: {first}")
    if last["Period"] != "2023M10" or abs(_decimal(last["Price"], context="Pink last") - Decimal("8.598018")) > Decimal("0.000001"):
        raise DataContractError(f"unexpected Pink Sheet last valid row: {last}")
    data = [
        {"기간": row["Period"], "명목가격": _json_number(_decimal(row["Price"], context=row["Period"]), 6)}
        for row in rows
    ]
    return {
        "id": "w15_pinksheet_nominal",
        "title": "국제 새우 명목가격 1960~2023 (달러/kg)",
        "subtitle": "World Bank Pink Sheet의 멕시코 서해안 냉동 새우 26~30미 규격 단일 프록시",
        "chartType": "line",
        "pillar": "S4",
        "telemetry": "STATIC",
        "syncDate": "World Bank Pink Sheet 2023년 10월",
        "source": "World Bank Pink Sheet, PinkSheet_Shrimp.csv",
        "sourceQuote": (
            f"{first['Period']},{first['Price']} | {last['Period']},{last['Price']} | "
            "2023M11 onward removed after constant 1079 corruption"
        ),
        "sit": "1960년 1월부터 2023년 10월까지 766개 월별 명목 관측치다. 단일 규격 계열이므로 전체 새우 가격으로 일반화하지 않는다.",
        "strat": "장기 가격선은 명목 단일 규격의 방향성을 보여줄 뿐이며, 인플레이션 조정 계열이나 전체 시장 대표가격으로 읽지 않아야 한다.",
        "xAxis": "기간",
        "series": [{"dataKey": "명목가격", "name": "명목가격", "color": "#14b8a6"}],
        "data": data,
        "unit": "달러/kg",
        "yUnit": "달러/kg",
    }


def _price_ladder_widget() -> dict[str, Any]:
    text = _read_text(PRICE_REPORT_PATH)
    crustaceans = text.split("CRUSTACEANS", 1)[1]
    whiteleg_block = crustaceans.split("Argentine red shrimp/", 1)[0]
    red_block = crustaceans.split("Argentine red shrimp/", 1)[1].split("Giant tiger prawn/", 1)[0]
    whiteleg_section = whiteleg_block.split("Whole", 1)[1].split("tails", 1)[0]
    whiteleg_prices = {}
    for size in ("20-30", "30-40", "40-50", "50-60"):
        match = _require_match(
            rf"^\s*{re.escape(size)}(?:\s+pc/kg)?\s+([\d.]+)\s+[\d.]+",
            whiteleg_section,
            label=f"Spain EXW whiteleg {size}",
            flags=re.MULTILINE,
        )
        whiteleg_prices[size] = _decimal(match.group(1), context=f"whiteleg {size}")
    red_section = red_block.split("Head-on, shell-on", 1)[1].split("Frozen on board", 1)[0]
    red_prices = {}
    for size in ("10-20", "20-30", "30-40", "40-60"):
        match = _require_match(
            rf"^.*?\b{re.escape(size)}(?:\s+pc/kg)?\s+([\d.]+)\s+[\d.]+",
            red_section,
            label=f"Spain EXW Argentine red {size}",
            flags=re.MULTILINE,
        )
        red_prices[size] = _decimal(match.group(1), context=f"red shrimp {size}")
    if "Spain EXW" not in whiteleg_section or "Spain EXW" not in red_section:
        raise DataContractError("Spain EXW condition was not found in both price blocks")
    data = [
        {"크기": "10-20", "양식": None, "자연산": _json_number(red_prices["10-20"], 2)},
        {"크기": "20-30", "양식": _json_number(whiteleg_prices["20-30"], 2), "자연산": _json_number(red_prices["20-30"], 2)},
        {"크기": "30-40", "양식": _json_number(whiteleg_prices["30-40"], 2), "자연산": _json_number(red_prices["30-40"], 2)},
        {"크기": "40-50/60", "양식": _json_number(whiteleg_prices["40-50"], 2), "자연산": _json_number(red_prices["40-60"], 2)},
        {"크기": "50-60", "양식": _json_number(whiteleg_prices["50-60"], 2), "자연산": None},
    ]
    quote_lines = [
        _normalize_space(line)
        for line in (whiteleg_section + "\n" + red_section).splitlines()
        if re.search(r"(?:10-20|20-30|30-40|40-50|40-60|50-60)", line)
    ]
    return {
        "id": "w16_spain_exw_ladder",
        "title": "규격별 호가 — 스페인 출고가 (유로/kg)",
        "subtitle": "2026년 6월 단일 시점 · Spain EXW 거래조건만 비교",
        "chartType": "bar",
        "pillar": "S4",
        "telemetry": "STATIC",
        "syncDate": "FAO GLOBEFISH 2026년 6월",
        "source": "FAO GLOBEFISH European Fish Price Report, June 2026",
        "sourceQuote": " | ".join(quote_lines),
        "sit": "Spain EXW에서 양식 흰다리새우는 크기가 작아질수록 6.35→5.30유로/kg로 내려가지만 자연산 붉은새우는 10~40미 구간 9.50유로/kg로 평탄하다.",
        "strat": "양식과 자연산은 같은 크기축에 놓여도 가격 형성 논리가 다르며, 거래조건이 다른 호가를 섞어 비교하지 않아야 한다.",
        "xKey": "크기",
        "bars": [
            {"key": "양식", "name": "양식 흰다리새우", "color": "#10b981"},
            {"key": "자연산", "name": "자연산 붉은새우", "color": "#f97316"},
        ],
        "data": data,
        "unit": "유로/kg",
        "yUnit": "유로/kg",
    }


def _cert_landscape_widget() -> dict[str, Any]:
    asc_path = ARCHIVE / "07_지속가능성·인증·ESG/ASC/20230701-Aquaculture_Stewardship_Council-ASC_Shrimp_Standard_v1.2.1.md"
    gdst_path = ARCHIVE / "07_지속가능성·인증·ESG/GDST/20250201-GDST-GDST_1.2_Core_Normative_Standards.md"
    msc_path = ARCHIVE / "07_지속가능성·인증·ESG/MSC/20260213-Marine_Stewardship_Council-MSC_Certification_of_Offshore_Argentine_Red_Shrimp.html"
    shaphari_path = ARCHIVE / "08_국가·산업리포트/India_MPEDA/legacy_archive/2026/LIST_OF_FARMS_CERTIFIED_FOR_SHAPHARI_040326.md"
    ssrt_path = ARCHIVE / "07_지속가능성·인증·ESG/legacy_shrimp_reports/Warmwater shrimp social risk profile — Vietnam ... - Seafood Watch.md"

    asc = _read_text(asc_path)
    gdst = _read_text(gdst_path)
    msc = _read_text(msc_path)
    shaphari = _read_text(shaphari_path)
    ssrt = _read_text(ssrt_path)
    asc_version = _require_match(r"Version\s+([\d.]+)", asc, label="ASC version").group(1)
    asc_date_text = _require_match(r"(\d{1,2}\s+July\s+2023)", asc, label="ASC document date").group(1)
    asc_date = datetime.strptime(asc_date_text, "%d %B %Y").date().isoformat()
    gdst_version = _require_match(r"Document Version\s+([\d.]+)", gdst, label="GDST version").group(1)
    gdst_date_text = _require_match(r"Document Date\s+([A-Za-z]+\s+\d{4})", gdst, label="GDST document date").group(1)
    gdst_date = datetime.strptime(gdst_date_text, "%B %Y").strftime("%Y-%m")
    msc_title = html.unescape(_normalize_space(_require_match(r"<title>\s*(.*?)\s*\|", msc, label="MSC title", flags=re.DOTALL).group(1)))
    msc_date_text = _normalize_space(_require_match(r"article-date[^>]*>\s*([^<]+)", msc, label="MSC date").group(1))
    msc_date = datetime.strptime(msc_date_text, "%d %B %Y").date().isoformat()
    if not shaphari.startswith("# LIST_OF_FARMS_CERTIFIED_FOR_SHAPHARI_040326"):
        raise DataContractError("SHAPHARI metadata title mismatch")
    ssrt_date_text = _require_match(
        r"Published\s+([A-Za-z]+\s+\d{2},\s+\d{4})\s+SEAFOOD SOCIAL RISK TOOL V(\d+)",
        ssrt,
        label="Seafood Watch publication metadata",
    )
    ssrt_date = datetime.strptime(ssrt_date_text.group(1), "%B %d, %Y").date().isoformat()
    custom_body = [
        {
            "name": "ASC 새우 표준",
            "issuer": "Aquaculture Stewardship Council",
            "version": asc_version,
            "date": asc_date,
            "scope": "전 세계 갑각류 농장 양식 생산 시스템과 지정 속의 양식 종",
        },
        {
            "name": "GDST 핵심 규범 표준",
            "issuer": "Global Dialogue on Seafood Traceability",
            "version": gdst_version,
            "date": gdst_date,
            "scope": "자연산·양식 수산물의 상호운용 이력추적 핵심 데이터와 추적 이벤트",
        },
        {
            "name": "MSC 인증",
            "issuer": "Marine Stewardship Council",
            "version": "MSC Fisheries Standard",
            "date": msc_date,
            "scope": "아르헨티나 근해 붉은새우 어업의 환경 지속가능성 인증",
        },
        {
            "name": "SHAPHARI",
            "issuer": "MPEDA",
            "version": "인증 승인 농장 목록",
            "date": "2026-03-04",
            "scope": "인도 SHAPHARI 인증 승인 새우 양식장 목록",
        },
        {
            "name": "수산물 사회위험 도구",
            "issuer": "Monterey Bay Aquarium Seafood Watch",
            "version": f"SSRT V{ssrt_date_text.group(2)}",
            "date": ssrt_date,
            "scope": "베트남 온수성 새우 양식·가공의 강제노동·인신매매·유해 아동노동 위험",
        },
    ]
    vintage_months = sorted(item["date"][:7] for item in custom_body)
    return {
        "id": "w21_cert_landscape",
        "title": "인증·표준 지형",
        "subtitle": "양식·자연산·이력추적·사회위험을 다루는 실존 문서 5종",
        "chartType": "none",
        "pillar": "S5",
        "telemetry": "STATIC",
        "syncDate": f"문서 빈티지 {vintage_months[0]}~{vintage_months[-1]}",
        "source": "ASC · GDST · MSC · MPEDA · Seafood Watch archived originals",
        "sourceQuote": (
            f"ASC Shrimp Standard Version {asc_version}, {asc_date_text}; "
            f"GDST Core Normative Standards Version {gdst_version}, {gdst_date_text}; "
            f"{msc_title}, {msc_date_text}; LIST OF FARMS APPROVED FOR SHAPHARI CERTIFICATION; "
            f"Warmwater shrimp social risk profile, {ssrt_date_text.group(1)}, SSRT V{ssrt_date_text.group(2)}"
        ),
        "sit": "다섯 문서는 같은 점수표가 아니라 양식장 운영, 자연산 어업, 데이터 이력추적, 농장 승인, 사회위험이라는 서로 다른 범위를 규율한다.",
        "strat": "인증 수나 커버리지 비율을 만들지 않고, 거래 단계와 위험 유형에 맞는 문서를 구분해 적용해야 한다.",
        "customBody": custom_body,
        "data": [],
    }


def _inherit_widgets(v3: Mapping[str, Any]) -> dict[str, dict[str, Any]]:
    # (pillar, syncDate, source 교정). source가 None이면 v3 문자열을 그대로 둔다.
    #
    # 가공생산 2건은 v3 출처가 두 가지로 낡았다. (1) 구버전 'FAOSTAT FishStatJ 2024.1.0'을
    # 인용하는데 페이지는 2026.1.0 위에 서 있고, (2) 'data/새우/' 경로는 2026-05-29
    # 13카테고리 재구성 때 Drive 아카이브로 옮겨져 레포에 존재하지 않는다.
    # 게다가 w03은 'Totals 행 교차 검증'이라 적었는데 그 Totals 행이 국가행과 합산되면
    # 전 연도 2배로 부풀려지는 함정(legacy shrimp_dashboard.json이 실제로 당한)이다.
    processing_source = (
        "FAO FishStat processed production, "
        "04_가공·완제품·기업/legacy_raw_data/9. 새우 가공 생산량 1976-2023.csv "
        "(Totals - Tonnes – net product weight 행 제외)"
    )
    specs = {
        "w03_processing": ("S2", "FAO 가공생산 2023년", processing_source),
        "w_proc1_type_production": ("S2", "FAO 가공생산 2023년", processing_source),
        "w_proc2_kr_import_type": ("S4", "KMI 2026년 1분기", None),
        "w50_kfas_bft_pathogen": ("S1", "2021년 서해권역 조사 (Gye et al., 2023)", None),
        "w_kr_shrimp_origin_price": ("S4", "KMI 2026년 1분기", None),
        "w_india_shaphari": ("S5", "MPEDA 2026-03-05", None),
        "w_vn_traceability_risk": ("S5", "Seafood Watch SSRT V2 (2026-01-07)", None),
    }
    source_by_id = {widget["id"]: widget for widget in v3["widgets"]}
    inherited: dict[str, dict[str, Any]] = {}
    for widget_id, (pillar, sync_date, source_override) in specs.items():
        if widget_id not in source_by_id:
            raise DataContractError(f"v3 inherited widget missing: {widget_id}")
        original = source_by_id[widget_id]
        widget = copy.deepcopy(original)
        original_data = copy.deepcopy(original.get("data"))
        telemetry = str(widget.get("telemetry", "")).upper()
        if telemetry not in ALLOWED_TELEMETRY:
            raise DataContractError(f"cannot normalize inherited telemetry for {widget_id}: {telemetry}")
        widget["telemetry"] = telemetry
        widget["syncDate"] = sync_date
        widget["pillar"] = pillar
        if source_override:
            widget["source"] = source_override
        if widget_id == "w03_processing":
            for key in ("sit", "situation"):
                if isinstance(widget.get(key), str):
                    widget[key] = widget[key].replace("85.3%", "87.9%")
        if widget.get("data") != original_data:
            raise DataContractError(f"inherited data array changed for {widget_id}")
        inherited[widget_id] = widget
    return inherited


def _build_kpis(
    widgets_by_id: Mapping[str, dict[str, Any]],
    globefish_facts: Mapping[str, Decimal],
    kcs_facts: Mapping[str, Decimal],
) -> dict[str, dict[str, Any]]:
    w01_last = widgets_by_id["w01_paradigm_shift"]["data"][-1]
    if w01_last["양식"] is None or w01_last["자연산"] is None:
        raise DataContractError("W01 2024 endpoint is missing")
    total = Decimal(str(w01_last["양식"])) + Decimal(str(w01_last["자연산"]))
    aquaculture_share = total and Decimal(str(w01_last["양식"])) / total * Decimal(100)
    whiteleg = next(
        row for row in widgets_by_id["w03_species_concentration"]["data"] if row["name"] == "흰다리새우"
    )
    trade_eok = globefish_facts["global_trade_value_billion"] * Decimal(10)
    ecuador_tonnes = globefish_facts["ecuador_2025_thousand_tonnes"] * Decimal(1000)
    korea_eok = kcs_facts["total_dollars"] / Decimal(100_000_000)
    korea_tonnes = kcs_facts["total_weight"] / Decimal(1000)
    return {
        "kpi1": {
            "title": "세계 새우 총생산",
            "value": f"{round(total):,}톤",
            "trend": "2024년",
            "desc": "양식과 자연산 합계",
            "telemetry": "STATIC",
            "syncDate": "FishStat 2026.1.0 · 2024년",
        },
        "kpi2": {
            "title": "양식 비중",
            "value": f"{float(aquaculture_share):.1f}%",
            "trend": "2024년",
            "desc": "총생산 중 양식",
            "telemetry": "STATIC",
            "syncDate": "FishStat 2026.1.0 · 2024년",
        },
        "kpi3": {
            "title": "흰다리새우 비중",
            "value": f"{whiteleg['value']:.1f}%",
            "trend": "2024년",
            "desc": "필터 후 종별 생산",
            "telemetry": "STATIC",
            "syncDate": "FishStat 2026.1.0 · 2024년",
        },
        "kpi4": {
            "title": "세계 교역액",
            "value": f"{float(trade_eok):.1f}억 달러",
            "trend": "2025년 수입액",
            "desc": "FAO GLOBEFISH 원문",
            "telemetry": "STATIC",
            "syncDate": "FAO GLOBEFISH · 2025년",
        },
        "kpi5": {
            "title": "최대 수출국 에콰도르",
            "value": f"{round(ecuador_tonnes):,}톤",
            "trend": f"{int(globefish_facts['ecuador_change']):+d}%",
            "desc": "2025년 수출량",
            "telemetry": "STATIC",
            "syncDate": "FAO GLOBEFISH · 2025년",
        },
        "kpi6": {
            "title": "한국 수입",
            "value": f"{float(korea_eok):.3f}억 달러",
            "trend": "5개월 누계",
            "desc": f"2026년 1~5월 누계 · {round(korea_tonnes):,}톤",
            "telemetry": "STATIC",
            "syncDate": "관세청 · 2026년 1~5월 누계",
        },
    }


def _nitemtrade_h1(hs_code: str) -> dict[str, dict[str, Decimal]]:
    rows = _read_csv_dicts(NITEMTRADE_PATH)
    totals: dict[str, dict[str, Decimal]] = defaultdict(
        lambda: {"imp_usd": Decimal(0), "imp_kg": Decimal(0)}
    )
    matched = 0
    for row in rows:
        if not row["year"].startswith("2026"):
            continue
        if row["hsCd"] != hs_code:
            continue
        country = row["statCdCntnKor1"]
        totals[country]["imp_usd"] += _decimal(row["impDlr"], context=f"nitem {hs_code} {country} usd")
        totals[country]["imp_kg"] += _decimal(row["impWgt"], context=f"nitem {hs_code} {country} kg")
        matched += 1
    if matched == 0:
        raise DataContractError(f"nitemtrade 2026 {hs_code} returned zero rows")
    return totals


def _series_production_2024(
    rows: Sequence[Mapping[str, str]],
) -> dict[str, dict[str, Decimal]]:
    totals: dict[str, dict[str, Decimal]] = defaultdict(
        lambda: {"total": Decimal(0), "aquaculture": Decimal(0)}
    )
    for row in rows:
        if row["PERIOD"] != "2024":
            continue
        country = row["COUNTRY.Name_En"]
        value = _decimal(row["VALUE"], context=f"FishStat 2024 {country}")
        totals[country]["total"] += value
        if row["PRODUCTION_SOURCE_DET.CODE"] in AQUACULTURE_SOURCES:
            totals[country]["aquaculture"] += value
    return totals


def _series_widgets(
    fish_rows: Sequence[Mapping[str, str]],
    cna_facts: Mapping[str, Decimal],
) -> tuple[dict[str, Any], dict[str, Any], dict[str, Any]]:
    raw = _nitemtrade_h1("030617")
    prepared = _nitemtrade_h1("160521")
    production = _series_production_2024(fish_rows)
    expected_raw_kg = {
        "베트남": Decimal("10604991"),
        "중국": Decimal("11568205"),
        "인도": Decimal("1830664"),
        "에콰도르": Decimal("1085664"),
        "태국": Decimal("828720"),
    }
    for country, expected in expected_raw_kg.items():
        actual = raw[SERIES_KCS_NAME[country]]["imp_kg"]
        if actual != expected:
            raise DataContractError(
                f"nitemtrade 030617 2026H1 {country} kg={actual}, expected {expected}"
            )

    window_data = []
    unit_data = []
    facts: dict[str, dict[str, Decimal]] = {}
    for country in SERIES_ORIGIN_ORDER:
        kcs_name = SERIES_KCS_NAME[country]
        raw_kg = raw[kcs_name]["imp_kg"]
        raw_usd = raw[kcs_name]["imp_usd"]
        prep_kg = prepared[kcs_name]["imp_kg"]
        unit = raw_usd / raw_kg
        facts[country] = {
            "raw_t": raw_kg / Decimal(1000),
            "prep_t": prep_kg / Decimal(1000),
            "unit": unit,
        }
        window_data.append(
            {
                "국가": country,
                "원물": _json_number(raw_kg / Decimal(1000), 1),
                "조제품": _json_number(prep_kg / Decimal(1000), 1),
            }
        )
        unit_data.append(
            {
                "국가": country,
                "단가": _json_number(unit, 2),
            }
        )

    korea_cna_pct = float(cna_facts["korea_share"] * 100)
    prod_line = {}
    for label, fao_name in SERIES_FAO_COUNTRIES.items():
        if fao_name not in production:
            raise DataContractError(f"FishStat 2024 missing {fao_name}")
        prod_line[label] = production[fao_name]

    vn = facts["베트남"]
    ec = facts["에콰도르"]
    india = facts["인도"]
    th = facts["태국"]
    cn = facts["중국"]
    sum_raw = sum((facts[name]["raw_t"] for name in SERIES_ORIGIN_ORDER), Decimal(0))
    sum_prep = sum((facts[name]["prep_t"] for name in SERIES_ORIGIN_ORDER), Decimal(0))

    roles = [
        {
            "name": "에콰도르",
            "role": "수출 표준 생산국",
            "korea": "한국 잔여 저단가 창구",
            "issuer": "수출 표준 생산국",
            "date": "CNA 2026년 1~5월 · 관세청 2026년 1~6월",
            "scope": (
                f"CNA 한국(COREA DEL SUR) 비중 {korea_cna_pct:.2f}%. "
                f"관세청 HS 030617 {float(ec['raw_t']):,.1f}톤, {float(ec['unit']):.2f}달러/kg. "
                "세계 수출 1위의 잔여 규격이다. SECA 발효·양허는 미확인."
            ),
        },
        {
            "name": "베트남",
            "role": "원물·조제품 이중 창구",
            "korea": "030617과 160521이 비슷한 무게",
            "issuer": "원물·조제품 이중 창구",
            "date": "FishStat 2024 · 관세청 2026년 1~6월",
            "scope": (
                f"필터 후 2024 생산 {int(round(prod_line['베트남']['total'])):,}톤. "
                f"한국 원물 {float(vn['raw_t']):,.1f}톤, 조제품 {float(vn['prep_t']):,.1f}톤. "
                "흰다리와 블랙타이거(*Penaeus monodon*)가 같이 있다. 한 원산지 이름으로 묶지 않는다."
            ),
        },
        {
            "name": "인도",
            "role": "미국·중국향 수출 생산국",
            "korea": "작은 원물 창구",
            "issuer": "미국·중국향 수출 생산국",
            "date": "FishStat 2024 · 관세청 2026년 1~6월",
            "scope": (
                f"필터 후 2024 생산 {int(round(prod_line['인도']['total'])):,}톤"
                f"(양식 {int(round(prod_line['인도']['aquaculture'])):,}). "
                f"한국 030617 {float(india['raw_t']):,.1f}톤, 조제품 {float(india['prep_t']):,.1f}톤. "
                "블랙타이거는 *Penaeus monodon*이다. 본진은 한국이 아니다."
            ),
        },
        {
            "name": "태국",
            "role": "원물 수입 가공국",
            "korea": "고단가 조제품 창구",
            "issuer": "원물 수입 가공국",
            "date": "관세청 2026년 1~6월",
            "scope": (
                f"한국 030617 {float(th['raw_t']):,.1f}톤({float(th['unit']):.2f}달러/kg), "
                f"조제품 {float(th['prep_t']):,.1f}톤. 원물보다 조제품이 더 무겁다. "
                "산지 완제품이 아니라 가공국 창구로 읽는다."
            ),
        },
        {
            "name": "중국",
            "role": "세계 1위 생산 · 한국 원물 1위",
            "korea": "030617 물량 1위, 조제품은 작다",
            "issuer": "세계 1위 생산 · 한국 원물 1위",
            "date": "FishStat 2024 · 관세청 2026년 1~6월",
            "scope": (
                f"필터 후 2024 생산 {int(round(prod_line['중국']['total'])):,}톤. "
                f"한국 030617 {float(cn['raw_t']):,.1f}톤({float(cn['unit']):.2f}달러/kg), "
                f"조제품 {float(cn['prep_t']):,.1f}톤. "
                "대하는 *Penaeus chinensis*만 가리킨다."
            ),
        },
        {
            "name": "한국",
            "role": "사서 쓰는 시장",
            "korea": "생산국이 아니라 창구 지도",
            "issuer": "사서 쓰는 시장",
            "date": "FishStat 2024 · 관세청 2026년 1~6월",
            "scope": (
                f"필터 후 2024 생산 {int(round(prod_line['한국']['total'])):,}톤. "
                f"시리즈 5개국에서 원물 {float(sum_raw):,.1f}톤, 조제품 {float(sum_prep):,.1f}톤을 사들인다. "
                "원산지마다 FTA가 다르고 SECA 발효 여부는 미확인이다."
            ),
        },
    ]

    w_roles = {
        "id": "w_series_country_roles",
        "title": "시리즈 6개국 역할",
        "subtitle": "에콰도르→베트남→인도→태국→중국→한국. 생산 순위와 한국 창구를 같은 표에 섞지 않음",
        "chartType": "none",
        "pillar": "S1",
        "telemetry": "STATIC",
        "syncDate": "FishStat 2024 · CNA 2026.1–5 · 관세청 2026.1–6",
        "source": "FAO FishStat 2026.1.0 · CNA May 2026 · 관세청 nitemtrade 2024-01~2026-06",
        "sourceQuote": (
            f"COREA DEL SUR={cna_facts['korea_share']}; "
            f"030617 kg VN={int(raw['베트남']['imp_kg'])} CN={int(raw['중국']['imp_kg'])} "
            f"IN={int(raw['인도']['imp_kg'])} EC={int(raw[SERIES_KCS_NAME['에콰도르']]['imp_kg'])} "
            f"TH={int(raw['태국']['imp_kg'])}"
        ),
        "sit": (
            "여섯 나라는 같은 새우 공급국이 아니다. 에콰도르는 수출 표준, 베트남은 원물·조제품 이중 창구, "
            "인도는 미국·중국향이 본진, 태국은 원물 수입 가공, 중국은 한국 원물 물량 1위, "
            "한국은 사서 쓰는 시장이다. 생산 톤은 FishStat 새우 필터 후 값이며 무역 제품중량과 빼지 않는다."
        ),
        "strat": (
            "원산지 이름으로 단가와 수율을 묶지 말고 세번(030617 vs 160521)과 학명을 계약서에 고정한다. "
            "한–에콰도르 SECA 특혜는 발효·양허를 확인하기 전 견적에 넣지 않는다."
        ),
        "customBody": roles,
        "unit": "역할 카드",
    }

    cheapest = min(SERIES_ORIGIN_ORDER, key=lambda name: facts[name]["unit"])
    dearest = max(SERIES_ORIGIN_ORDER, key=lambda name: facts[name]["unit"])
    w_windows = {
        "id": "w_series_kr_windows",
        "title": "한국 창구 물량 (톤)",
        "subtitle": "2026년 1~6월 · HS 030617 원물과 160521 조제품 · 연환산 금지",
        "chartType": "bar",
        "pillar": "S3",
        "telemetry": "SYNCED",
        "syncDate": "관세청 2026년 1~6월",
        "source": "관세청 nitemtrade HS×상대국 2024-01~2026-06",
        "sourceQuote": (
            f"year prefix 2026; hsCd 030617/160521; "
            f"VN raw={int(raw['베트남']['imp_kg'])} prep={int(prepared['베트남']['imp_kg'])}"
        ),
        "sit": (
            f"원물 물량은 중국 {float(cn['raw_t']):,.1f}톤, 베트남 {float(vn['raw_t']):,.1f}톤이다. "
            f"조제품은 베트남 {float(vn['prep_t']):,.1f}톤, 태국 {float(th['prep_t']):,.1f}톤이다. "
            f"인도 원물은 {float(india['raw_t']):,.1f}톤, 에콰도르는 {float(ec['raw_t']):,.1f}톤이다. "
            "같은 원산지라도 세번이 갈리면 창구가 다르다."
        ),
        "strat": "베트남을 싼 원물로, 에콰도르를 한국 주력으로 가정하지 말고 세번별 물량으로 발주 창구를 나눈다.",
        "xKey": "국가",
        "bars": [
            {"key": "원물", "name": "030617 원물", "color": "#38bdf8"},
            {"key": "조제품", "name": "160521 조제품", "color": "#f59e0b"},
        ],
        "data": window_data,
        "unit": "톤",
        "yUnit": "톤",
    }
    w_unit = {
        "id": "w_series_kr_unit",
        "title": "한국 창구 단가 (달러/kg)",
        "subtitle": "2026년 1~6월 HS 030617 신고액÷중량 · 조제품 단가와 섞지 않음",
        "chartType": "bar",
        "pillar": "S4",
        "telemetry": "SYNCED",
        "syncDate": "관세청 2026년 1~6월",
        "source": "관세청 nitemtrade HS×상대국 2024-01~2026-06",
        "sourceQuote": (
            f"EC {float(ec['unit']):.4f} · IN {float(india['unit']):.4f} · "
            f"CN {float(cn['unit']):.4f} · VN {float(vn['unit']):.4f} · "
            f"TH {float(th['unit']):.4f} USD/kg"
        ),
        "sit": (
            f"030617 단가는 {cheapest} {float(facts[cheapest]['unit']):.2f}달러/kg가 가장 낮고 "
            f"{dearest} {float(facts[dearest]['unit']):.2f}달러/kg가 가장 높다. "
            f"중국은 물량 1위지만 단가는 {float(cn['unit']):.2f}달러로 중간이다."
        ),
        "strat": "낮은 단가를 수율로 바꾸지 말고, 글레이즈·카운트·잔류 시험을 같은 로트에 붙인 뒤에만 비교한다.",
        "xKey": "국가",
        "bars": [{"key": "단가", "name": "030617 단가", "color": "#14b8a6"}],
        "data": unit_data,
        "unit": "달러/kg",
        "yUnit": "달러/kg",
    }
    return w_roles, w_windows, w_unit


def build_payload() -> tuple[dict[str, Any], dict[str, Decimal]]:
    _require_file(V3_PATH)
    v3 = json.loads(V3_PATH.read_text(encoding="utf-8"))
    if not isinstance(v3, dict) or not isinstance(v3.get("widgets"), list):
        raise DataContractError("v3 JSON shape is invalid")

    fish_rows, metrics = _load_fishstat()
    metrics["sofia_aquaculture"] = _parse_sofia_aquaculture()
    w01, w02, w03 = _fishstat_widgets(fish_rows, metrics)
    w04 = _argentina_widget()
    w08 = _processing_reversal_widget()
    w09 = _avanti_widget()
    w10, w14, globefish_facts = _globefish_widgets()
    w11, cna_facts = _cna_widget()
    w12 = _infofish_widget()
    w13, kcs_facts = _kcs_widget()
    w_roles, w_windows, w_unit = _series_widgets(fish_rows, cna_facts)
    w15 = _pinksheet_widget()
    w16 = _price_ladder_widget()
    w21 = _cert_landscape_widget()
    inherited = _inherit_widgets(v3)

    widgets = [
        w01,
        w02,
        w03,
        w_roles,
        w04,
        inherited["w50_kfas_bft_pathogen"],
        inherited["w03_processing"],
        inherited["w_proc1_type_production"],
        w08,
        w09,
        w10,
        w11,
        w12,
        w13,
        w_windows,
        w14,
        w15,
        w16,
        inherited["w_kr_shrimp_origin_price"],
        w_unit,
        inherited["w_proc2_kr_import_type"],
        inherited["w_india_shaphari"],
        inherited["w_vn_traceability_risk"],
        w21,
    ]
    by_id = {widget["id"]: widget for widget in widgets}
    if len(by_id) != len(widgets):
        raise DataContractError("duplicate widget id in v4 payload")
    payload = {
        "kpis": _build_kpis(by_id, globefish_facts, kcs_facts),
        "widgets": widgets,
    }
    return payload, metrics


def validate_payload(payload: Mapping[str, Any]) -> list[str]:
    errors: list[str] = []
    widgets = payload.get("widgets")
    kpis = payload.get("kpis")
    if not isinstance(widgets, list):
        return ["widgets is not a list"]
    if not isinstance(kpis, dict):
        return ["kpis is not an object"]
    if len(widgets) != EXPECTED_WIDGET_COUNT:
        errors.append(f"widget count={len(widgets)}, expected {EXPECTED_WIDGET_COUNT}")
    if len(kpis) != 6:
        errors.append(f"KPI count={len(kpis)}, expected 6")

    pillar_counts: Counter[str] = Counter()
    for widget in widgets:
        widget_id = widget.get("id", "<missing id>")
        pillar = widget.get("pillar")
        if pillar not in ALLOWED_PILLARS:
            errors.append(f"{widget_id}: invalid or missing pillar={pillar!r}")
        else:
            pillar_counts[pillar] += 1
        for field in ("source", "syncDate", "telemetry"):
            if not widget.get(field):
                errors.append(f"{widget_id}: missing {field}")
        if widget.get("telemetry") not in ALLOWED_TELEMETRY:
            errors.append(f"{widget_id}: telemetry={widget.get('telemetry')!r}")
        if widget.get("chartType") == "line":
            if not widget.get("xAxis") or not widget.get("series"):
                errors.append(f"{widget_id}: line chart lacks xAxis/series")
            forbidden = [key for key in ("xKey", "bars", "lines", "areas") if key in widget]
            if forbidden:
                errors.append(f"{widget_id}: line chart has NEW FORMAT keys {forbidden}")
    if dict(sorted(pillar_counts.items())) != EXPECTED_PILLARS:
        errors.append(
            f"pillar distribution={dict(sorted(pillar_counts.items()))}, expected={EXPECTED_PILLARS}"
        )

    serialized = json.dumps(payload, ensure_ascii=False, sort_keys=True)
    for banned in BANNED_STRINGS:
        if banned in serialized:
            errors.append(f"banned string found: {banned}")
    compact = re.sub(r"[,_\s.]", "", serialized)
    for value in UNFILTERED_VALUES:
        if value in compact:
            errors.append(f"unfiltered FishStat value found: {value}")
    return errors


def _format_decimal(value: Decimal) -> str:
    if value == value.to_integral_value():
        return f"{int(value):,}"
    return f"{value:,.3f}".rstrip("0").rstrip(".")


def print_gate_results(results: Sequence[GateResult]) -> None:
    for index, result in enumerate(results, start=1):
        status = "PASS" if result.passed else "FAIL"
        print(
            f"assert {index}: {status} | {result.name} | "
            f"actual={_format_decimal(result.actual)} | "
            f"expected={_format_decimal(result.expected)} | "
            f"tolerance=±{_format_decimal(result.tolerance)}"
        )


def run_verify() -> int:
    try:
        payload, metrics = build_payload()
        results = canonical_gate_results(metrics)
        print_gate_results(results)
        integrity_errors = validate_payload(payload)
        if integrity_errors:
            for error in integrity_errors:
                print(f"integrity: FAIL | {error}")
        else:
            pillars = ",".join(f"{key}:{EXPECTED_PILLARS[key]}" for key in sorted(EXPECTED_PILLARS))
            print(
                f"integrity: PASS | widgets={EXPECTED_WIDGET_COUNT} | kpis=6 | "
                f"pillars={pillars} | LIVE=0"
            )
        return 0 if all(result.passed for result in results) and not integrity_errors else 1
    except Exception as exc:
        print(f"verify: FAIL | {type(exc).__name__}: {exc}", file=sys.stderr)
        return 1


def write_payload(payload: Mapping[str, Any]) -> None:
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    serialized = json.dumps(payload, ensure_ascii=False, indent=2) + "\n"
    temporary_path: Path | None = None
    try:
        with tempfile.NamedTemporaryFile(
            mode="w",
            encoding="utf-8",
            dir=OUTPUT_PATH.parent,
            prefix=f".{OUTPUT_PATH.name}.",
            suffix=".tmp",
            delete=False,
        ) as stream:
            temporary_path = Path(stream.name)
            stream.write(serialized)
            stream.flush()
            os.fsync(stream.fileno())
        os.replace(temporary_path, OUTPUT_PATH)
        temporary_path = None
    finally:
        if temporary_path is not None and temporary_path.exists():
            temporary_path.unlink()


def generate() -> int:
    payload, metrics = build_payload()
    results = canonical_gate_results(metrics)
    failed = [result for result in results if not result.passed]
    if failed:
        print_gate_results(results)
        raise DataContractError("canonical assertions failed; output was not written")
    errors = validate_payload(payload)
    if errors:
        raise DataContractError("payload integrity failed: " + "; ".join(errors))
    write_payload(payload)
    counts = Counter(widget["pillar"] for widget in payload["widgets"])
    print(f"wrote {OUTPUT_PATH}")
    print(f"widgets={len(payload['widgets'])} kpis={len(payload['kpis'])}")
    print("pillars=" + ",".join(f"{pillar}:{counts[pillar]}" for pillar in sorted(counts)))
    return 0


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--verify", action="store_true", help="run all gates without writing JSON")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    if args.verify:
        return run_verify()
    try:
        return generate()
    except Exception as exc:
        print(f"generation: FAIL | {type(exc).__name__}: {exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    sys.exit(main())
