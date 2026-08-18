#!/usr/bin/env python3
"""파노피 추정실적 xlsx -> public/data/panofi/panofi_actuals.json

최신 「2. 추정실적 (2026년 N월).xlsx」 한 장이 1~N월 누계를 담는다.
월별 파일을 잇거나 부분연을 연환산하지 않는다.

뽑는 것:
  0) 실적 시트 좌측 누계 블록(판매/생산) + 전년 동기 블록 + 원장 BEP
  1) 월별 손익 (실적 시트 N열부터)
  2) 연도별 손익 (실적 시트 연도 블록)
  3) 척별 손익·원가 (실적(생산))
  4) 척별×어종×사이즈 생산량 (매출단가)

작성자 주석대로 사실상 연 결산이라 월별 원가 배분 변동성이 크다.
5월 매출원가 음수는 이월 정산이지 실제 마이너스 원가가 아니다.
"""
from __future__ import annotations

import json
import os
import re
import sys
import unicodedata
from hashlib import sha256
from pathlib import Path

try:
    import openpyxl
except ImportError:  # pragma: no cover
    print("openpyxl 필요: pip install openpyxl", file=sys.stderr)
    raise SystemExit(2)

OUT = Path(__file__).resolve().parents[1] / "public/data/panofi/panofi_actuals.json"

DRIVE_PANOFI = Path.home() / (
    "Library/CloudStorage/GoogleDrive-cutekorea@gmail.com/내 드라이브/"
    "신라그룹/11_Panofi_Cosmo_GGL /11. PANOFI/Panofi"
)
LEGACY_PANOFI = Path.home() / "my-project/11_Panofi_Cosmo_GGL /11. PANOFI/Panofi"

VESSELS = ["P/MAS", "P/DIS", "P/FORE", "P/PATH", "P/COMM", "P/QUE", "P/GRA"]
VESSEL_KO = {
    "P/MAS": "마스터", "P/DIS": "디스커버러", "P/FORE": "포러너", "P/PATH": "패스파인더",
    "P/COMM": "커맨더", "P/QUE": "퀸", "P/GRA": "그레이스",
}
SPECIES_KO = {"YF": "황다랑어", "SJ": "가다랑어", "BE": "눈다랑어", "잡어": "잡어"}


def nfc(s: str) -> str:
    return unicodedata.normalize("NFC", s)


def num(v):
    """숫자만 통과시킨다. 엑셀 문자열·수식잔재는 None 으로 떨군다."""
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        return round(float(v), 2)
    return None


def find_src() -> Path:
    env = os.environ.get("PANOFI_ACTUALS_XLSX")
    if env:
        p = Path(env)
        if not p.exists():
            raise FileNotFoundError(p)
        return p

    cands: list[tuple[int, Path]] = []
    for root in (DRIVE_PANOFI, LEGACY_PANOFI):
        if not root.exists():
            continue
        for p in root.rglob("*.xlsx"):
            name = nfc(p.name)
            if name.startswith("~$"):
                continue
            if "추정실적" not in name:
                continue
            m = re.search(r"2026년\s*(\d+)월", name)
            if not m:
                continue
            cands.append((int(m.group(1)), p))
    if not cands:
        raise FileNotFoundError("2026 추정실적 xlsx 없음")
    cands.sort(key=lambda x: x[0])
    return cands[-1][1]


def monthly(ws) -> list[dict]:
    """실적 시트 r4=월 라벨, r5~r22=계정. N열(14)부터 12개월."""
    accounts = {
        5: "수량MT", 6: "평균단가", 7: "매출액", 8: "매출원가", 9: "원가율",
        10: "기초재고액", 11: "당기생산액", 13: "당기판매액", 14: "기말재고액",
        15: "매출총이익", 16: "판매관리비", 17: "영업이익", 18: "기타수익비용",
        19: "금융비용", 21: "법인세비용", 22: "당기순이익",
    }
    out = []
    for i in range(12):
        col = 14 + i
        label = ws.cell(row=4, column=col).value
        rec = {"month": label, "monthIndex": i + 1}
        if any(num(ws.cell(row=r, column=col).value) is not None for r in accounts):
            for r, key in accounts.items():
                rec[key] = num(ws.cell(row=r, column=col).value)
            out.append(rec)
    return out


def annual(ws) -> list[dict]:
    """실적 시트 r30=연도 라벨, r31~r40. N열(14)부터."""
    accounts = {
        31: "수량MT", 32: "평균단가", 33: "매출액", 34: "매출원가", 35: "원가율",
        36: "기초재고액", 37: "당기생산액", 39: "당기판매액", 40: "기말재고액",
    }
    out = []
    for i in range(8):
        col = 14 + i
        label = ws.cell(row=30, column=col).value
        if not label:
            continue
        rec = {"year": str(label).replace("년", "")}
        for r, key in accounts.items():
            rec[key] = num(ws.cell(row=r, column=col).value)
        if rec["수량MT"] is not None:
            out.append(rec)
    return out


def ledger_side(ws, qty_row: int) -> dict:
    """판매기준(B) / 생산기준(C) 한 블록. qty_row 는 수량 행."""
    def col(c, off):
        return num(ws.cell(row=qty_row + off, column=c).value)

    return {
        "수량MT": col(2, 0),
        "평균단가": col(2, 1),
        "매출액": col(2, 2),
        "매출원가": col(2, 3),
        "원가율": col(2, 4),
        "기초재고액": col(2, 5),
        "당기생산액": col(2, 6),
        "당기판매액": col(2, 8),
        "기말재고액": col(2, 9),
        "기말재고MT": col(4, 9),
        "매출총이익": col(2, 10),
        "판매관리비": col(2, 11),
        "영업이익": col(2, 12),
        "기타수익비용": col(2, 13),
        "금융비용": col(2, 14),
        "법인세비용": col(2, 16),
        "당기순이익": col(2, 17),
        "bep어가": col(2, 19),
        "생산수량MT": col(3, 0),
        "생산평균단가": col(3, 1),
        "생산매출액": col(3, 2),
        "생산제조원가": col(3, 3),
        "생산원가율": col(3, 4),
        "생산영업이익": col(3, 12),
        "생산당기순이익": col(3, 17),
        "생산bep어가": col(3, 19),
    }


def summary(ws) -> dict:
    """좌측 올해 누계(2행 표지)와 전년 동기(28행 표지)."""
    now_title = ws.cell(row=2, column=1).value
    prior_title = ws.cell(row=28, column=1).value
    months = num(ws.cell(row=2, column=7).value)
    return {
        "periodLabel": str(now_title) if now_title else None,
        "months": int(months) if months else None,
        "sales": ledger_side(ws, 5),
        "priorPeriodLabel": str(prior_title) if prior_title else None,
        "prior": ledger_side(ws, 31),
    }


def by_vessel(ws) -> dict:
    """실적(생산) 시트. 척은 E~K열(5~11), 합계 L열(12)."""
    summary_rows = {
        5: "생산량MT", 6: "생산매출액", 7: "제조원가", 8: "생산총이익",
        9: "판매비및관리비", 10: "영업이익", 11: "기타수익비용", 12: "세전이익",
    }
    cost_rows = {
        18: ("재료비", "연료비"), 19: ("재료비", "윤활유비"),
        21: ("노무비", "급여및임금"), 22: ("노무비", "상여및수당"),
        23: ("노무비", "식료품비"), 25: ("노무비", "선원의료비"), 26: ("노무비", "잡급"),
        28: ("경비", "선용품비"), 29: ("경비", "운반비"), 30: ("경비", "감가상각비"),
        31: ("경비", "어구비"), 32: ("경비", "수선비"), 33: ("경비", "세금과공과"),
        34: ("경비", "선박보험료"), 35: ("경비", "선원보험료"), 36: ("경비", "여비교통비"),
        37: ("경비", "복리후생비"), 38: ("경비", "통신비"), 39: ("경비", "보관비"),
        40: ("경비", "입어료"), 41: ("경비", "항만비"), 42: ("경비", "경비료"),
        43: ("경비", "지급수수료"),
    }

    vessels = []
    for vi, code in enumerate(VESSELS):
        col = 5 + vi
        rec = {"code": code, "name": VESSEL_KO[code]}
        for r, key in summary_rows.items():
            rec[key] = num(ws.cell(row=r, column=col).value)
        costs = []
        for r, (group, item) in cost_rows.items():
            v = num(ws.cell(row=r, column=col).value)
            if v:
                costs.append({"group": group, "item": item, "usd": v})
        rec["costs"] = costs
        vessels.append(rec)

    totals = {}
    for r, key in summary_rows.items():
        totals[key] = num(ws.cell(row=r, column=12).value)
    cost_totals = []
    for r, (group, item) in cost_rows.items():
        v = num(ws.cell(row=r, column=12).value)
        if v:
            cost_totals.append({"group": group, "item": item, "usd": v})

    return {"vessels": vessels, "totals": totals, "costTotals": cost_totals}


def catch_mix(ws) -> list[dict]:
    """매출단가 시트 r5~r30. 어종(C열) x 사이즈(D열) x 척(G~M열) x 합계(N=14)."""
    rows = []
    species = None
    for ri in range(5, 32):
        sp = ws.cell(row=ri, column=3).value
        size = ws.cell(row=ri, column=4).value
        if sp:
            species = str(sp).strip()
        if not size:
            continue
        size = str(size).strip()
        if size in ("계",):
            continue
        total = num(ws.cell(row=ri, column=14).value)
        if not total:
            continue
        per = {}
        for vi, code in enumerate(VESSELS):
            per[code] = num(ws.cell(row=ri, column=7 + vi).value) or 0
        rows.append({
            "speciesCode": species,
            "species": SPECIES_KO.get(species, species),
            "size": size,
            "byVessel": per,
            "totalMT": total,
        })
    return rows


def main() -> int:
    try:
        src = find_src()
    except FileNotFoundError as exc:
        print(f"원자료 없음: {exc}", file=sys.stderr)
        return 1

    wb = openpyxl.load_workbook(src, data_only=True)
    ws_act, ws_prod, ws_price = wb["실적"], wb["실적(생산)"], wb["매출단가"]

    months = monthly(ws_act)
    years = annual(ws_act)
    vessels = by_vessel(ws_prod)
    mix = catch_mix(ws_price)
    ytd = summary(ws_act)

    v_sum = sum(v["생산량MT"] or 0 for v in vessels["vessels"])
    v_tot = vessels["totals"]["생산량MT"] or 0
    drift = round(v_sum - v_tot, 1)
    print(f"척별 생산량 합 {v_sum:,.1f} vs 합계 {v_tot:,.1f} (차 {drift})", file=sys.stderr)

    mix_sum = round(sum(r["totalMT"] for r in mix), 1)
    print(f"어종·사이즈 생산량 합 {mix_sum:,.1f}톤 · {len(mix)}행", file=sys.stderr)
    print(
        f"월별 {len(months)}개월 · 연도별 {len(years)}개년 · 척 {len(vessels['vessels'])} · {ytd['periodLabel']}",
        file=sys.stderr,
    )

    month_n = ytd["months"] or len(months)
    payload = {
        "meta": {
            "source": nfc(src.name),
            "sha256": sha256(src.read_bytes()).hexdigest(),
            "basis": f"2026년 1~{month_n}월 누계 (판매기준·생산기준 병기)",
            "syncDate": "2026-08-17",
            "caveat": "작성자 주석대로 사실상 연 결산이라 월별 원가 배분 변동성이 크다. "
                      "5월은 판매 342톤에 매출원가가 음수로 잡히는데 이월 정산의 결과지 "
                      "실제 마이너스 원가가 아니다. 월별은 참고로만 보고 판단은 누계로 한다. "
                      "부분연을 연환산하지 않는다.",
            "vesselProductionDrift": drift,
            "catchMixTotalMT": mix_sum,
        },
        "summary": ytd,
        "monthly": months,
        "annual": years,
        "byVessel": vessels,
        "catchMix": mix,
    }

    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(payload, ensure_ascii=False, indent=1), encoding="utf8")
    print(f"-> {OUT} ({OUT.stat().st_size // 1024}KB)", file=sys.stderr)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
