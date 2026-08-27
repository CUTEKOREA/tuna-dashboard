#!/usr/bin/env python3
"""Convert one TTA reefer movement workbook into a checked weekly JSON snapshot."""

from __future__ import annotations

import argparse
from datetime import date
import json
import math
from pathlib import Path
import re
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string


WEEK_RE = re.compile(r'week\s+(\d+)', re.IGNORECASE)
PERIOD_RE = re.compile(r'(\d{2})/(\d{2})/(\d{2})\s*-\s*(\d{2})/(\d{2})/(\d{2})')
ERROR_VALUES = {'#REF!', '#DIV/0!', '#VALUE!', '#N/A', '#NAME?', '#NUM!', '#NULL!'}
EXPECTED_DESTINATION_HEADERS = {
    'F': 'ASIAN', 'G': 'AEC', 'H': 'AYA', 'I': 'CMC', 'J': 'DIA', 'K': 'GB',
    'L': 'GPZ', 'M': 'ISA', 'N': 'I-TAIL', 'O': 'KF', 'P': 'MMP', 'Q': 'PCI',
    'R': 'FOOD', 'S': 'POP', 'T': 'PTY', 'U': 'RMK', 'V': 'RS', 'W': 'SK',
    'X': 'SIF', 'Y': 'SPA', 'Z': 'SCC', 'AA': 'SE', 'AB': 'SEAP', 'AC': 'TCC',
    'AD': 'TOV', 'AE': 'TUG', 'AF': 'TUM', 'AG': 'UC', 'AH': 'SHIP',
}


def number(value: Any) -> float | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        result = float(value)
        return result if math.isfinite(result) else None
    if isinstance(value, str):
        normalized = value.strip().replace(',', '')
        if not normalized or normalized == '-':
            return None
        try:
            result = float(normalized)
        except ValueError:
            return None
        return result if math.isfinite(result) else None
    return None


def display(value: Any) -> str:
    numeric = number(value)
    if numeric is None:
        return str(value).strip() if value is not None else ''
    if numeric.is_integer():
        return f'{int(numeric):,}'
    return f'{numeric:,.3f}'.rstrip('0').rstrip('.')


def period_date(day: str, month: str, year: str) -> str:
    return date(2000 + int(year), int(month), int(day)).isoformat()


def parse_workbook(path: Path) -> tuple[int, str, str, list[dict[str, Any]], float]:
    match = WEEK_RE.search(path.name)
    if not match:
        raise ValueError(f'파일명에서 주차를 찾을 수 없습니다: {path.name}')
    week = int(match.group(1))

    formulas = load_workbook(path, data_only=False, read_only=False)
    values = load_workbook(path, data_only=True, read_only=False)
    expected_sheet = f'WEEK {week}'
    if values.sheetnames != [expected_sheet]:
        raise ValueError(f'예상 시트 {expected_sheet!r}와 다릅니다: {values.sheetnames!r}')
    source_sheet = formulas[expected_sheet]
    sheet = values[expected_sheet]

    errors: list[str] = []
    for row in sheet.iter_rows():
        for cell in row:
            value = cell.value
            if cell.data_type == 'e' or (isinstance(value, str) and value in ERROR_VALUES):
                errors.append(f'{cell.coordinate}={value}')
    if errors:
        raise ValueError('엑셀 수식 오류: ' + ', '.join(errors))

    if number(sheet['AJ2'].value) != week:
        raise ValueError(f'파일명 주차와 AJ2가 다릅니다: {week} / {sheet["AJ2"].value!r}')
    period_match = PERIOD_RE.search(str(sheet['A1'].value))
    if not period_match:
        raise ValueError(f'A1에서 보고기간을 찾을 수 없습니다: {sheet["A1"].value!r}')
    start_date = period_date(*period_match.groups()[:3])
    end_date = period_date(*period_match.groups()[3:])

    headers: dict[int, str] = {}
    for column_name, expected in EXPECTED_DESTINATION_HEADERS.items():
        column = column_index_from_string(column_name)
        actual = str(sheet.cell(6, column).value or '').strip()
        if actual != expected:
            raise ValueError(
                f'템플릿 배분 헤더 불일치: {expected_sheet}!{column_name}6={actual!r}, 예상 {expected!r}',
            )
        headers[column] = actual
    rows: list[dict[str, Any]] = []
    grand_total = 0.0
    for row in range(7, sheet.max_row + 1):
        carrier = sheet.cell(row, 1).value
        if isinstance(carrier, str) and 'SONGKHLA PORT' in carrier.upper():
            break
        berthing = sheet.cell(row, 2).value
        if not carrier or not berthing:
            continue

        deliveries: dict[str, str] = {}
        allocated_total = 0.0
        for column, header in headers.items():
            value = sheet.cell(row, column).value
            formula_value = source_sheet.cell(row, column).value
            if value is None:
                if isinstance(formula_value, str) and formula_value.startswith('='):
                    raise ValueError(f'수식 캐시 누락: {expected_sheet}!{sheet.cell(row, column).coordinate}')
                continue
            deliveries[header] = display(value)
            if header != 'SHIP':
                numeric = number(value)
                if numeric is None:
                    raise ValueError(f'배분량이 숫자가 아닙니다: {expected_sheet}!{sheet.cell(row, column).coordinate}')
                allocated_total += numeric

        source_total = number(sheet.cell(row, column_index_from_string('AI')).value)
        if source_total is None:
            raise ValueError(f'선박 총량이 없습니다: {expected_sheet}!AI{row}')
        if abs(source_total - allocated_total) > 0.001:
            raise ValueError(
                f'선박 배분 합계 불일치: {carrier} 원문 {source_total} / 재계산 {allocated_total}',
            )
        remark = sheet.cell(row, column_index_from_string('AJ')).value
        if remark not in (None, ''):
            deliveries['OTHER'] = display(remark)
        rows.append({
            'carrier': str(carrier).strip(),
            'date': display(berthing),
            'status': '주간 보고 기록',
            'daysRemaining': None,
            'priority': '이력',
            'deliveries': deliveries,
        })
        grand_total += allocated_total

    if not rows:
        raise ValueError('방콕항 운반선 행이 없습니다')
    return week, start_date, end_date, rows, grand_total


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument('workbook', type=Path)
    parser.add_argument('--output', type=Path)
    parser.add_argument('--check', action='store_true')
    args = parser.parse_args()

    workbook = args.workbook.expanduser().resolve()
    if not workbook.is_file():
        raise FileNotFoundError(workbook)
    week, start_date, end_date, rows, grand_total = parse_workbook(workbook)
    output = (args.output or Path(f'data/reefer_week{week}.json')).expanduser().resolve()
    content = json.dumps(rows, ensure_ascii=False, indent=2) + '\n'

    summary = {
        'week': week,
        'period': [start_date, end_date],
        'vessels': len(rows),
        'grandTotalMt': round(grand_total, 3),
        'output': str(output),
    }
    print(json.dumps(summary, ensure_ascii=False, indent=2))
    if args.check:
        if not output.is_file() or output.read_text(encoding='utf-8') != content:
            raise ValueError(f'출력 JSON이 현재 원문과 다릅니다: {output}')
        return 0
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(content, encoding='utf-8')
    return 0


if __name__ == '__main__':
    raise SystemExit(main())
